"""
Individual Sheet Processing - NEW TEMPLATE LAYOUT
Months as columns (C-N), KPIs as rows (4-8)

This replaces the old update_individual_sheet() function in function_app_local.py
for use with the redesigned templates where months are columns instead of rows.
"""

import logging
import re
from datetime import date
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION - NEW TEMPLATE LAYOUT
# =============================================================================

# Month to Column mapping (NEW - months are columns)
MONTH_COLUMNS = {
    'Jan': 3, 'January': 3,    # Column C
    'Feb': 4, 'February': 4,   # Column D
    'Mar': 5, 'March': 5,      # Column E
    'Apr': 6, 'April': 6,      # Column F
    'May': 7,                  # Column G
    'Jun': 8, 'June': 8,       # Column H
    'Jul': 9, 'July': 9,       # Column I
    'Aug': 10, 'August': 10,   # Column J
    'Sep': 11, 'Sept': 11, 'September': 11,  # Column K
    'Oct': 12, 'October': 12,  # Column L
    'Nov': 13, 'November': 13, # Column M
    'Dec': 14, 'December': 14  # Column N
}

# KPI to Row mapping - Physio (NEW - KPIs are rows)
KPI_ROWS_PHYSIO = {
    'BillingsKPI': 4,
    'Ceased Services': 5,
    'Documentation': 6,
    'Admin': 7,
    'Attitude': 8
}

# KPI to Row mapping - OT (NEW - KPIs are rows)
KPI_ROWS_OT = {
    'BillingsKPI': 4,
    'Compliance': 5,
    'Referrer Engagement': 6,
    'Capacity': 7,
    'Attitude': 8
}

# Threshold display locations in new template
THRESHOLD_ROWS = {
    'billings': {
        'Grad': 12,     # Row 12, columns C-E
        'CA': 13,       # Row 13, columns C-E
        'Senior': 14    # Row 14, columns C-E
    },
    'ceased': {
        'excellent': 17,  # Row 17, column D
        'good': 18,       # Row 18, column D
        'below': 19       # Row 19, column D
    }
}

# Data range for conditional formatting (C-O = columns 3-15, includes Average)
DATA_START_COL = 3   # Column C
DATA_END_COL = 15    # Column O (Average)

DEFAULT_COLORS = {
    'red': 'FFE47373',
    'amber': 'FFFFB74D',
    'yellow': 'FFFFF176',
    'green': 'FF81C784',
    'blue': 'FF4FC3F7',
    'white': 'FFFFFFFF'
}

# Month name to number mapping for competency history
MONTH_NAME_TO_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'June': 6,
    'July': 7, 'Aug': 8, 'Sept': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

# Reverse mapping: column number to month name
COL_TO_MONTH = {
    3: 'Jan', 4: 'Feb', 5: 'Mar', 6: 'Apr', 7: 'May', 8: 'June',
    9: 'July', 10: 'Aug', 11: 'Sept', 12: 'Oct', 13: 'Nov', 14: 'Dec'
}


def get_competency_for_month(therapist_name, month_name, year, config):
    """
    Get the competency that was active for a therapist in a specific month.
    
    Args:
        therapist_name: Name of the therapist
        month_name: Month name ('Jan', 'Feb', etc.)
        year: Year (e.g., 2026)
        config: Config dict containing 'competency_history' and 'therapists'
        
    Returns:
        str: Competency level ('Grad', 'CA', 'Senior') or None if not found
    """
    month_num = MONTH_NAME_TO_NUM.get(month_name)
    if not month_num:
        return None
    
    target_date = date(year, month_num, 15)
    
    history = config.get('competency_history', [])
    therapist_records = [
        r for r in history 
        if r.get('Name', '').strip().lower() == therapist_name.strip().lower()
    ]
    
    if therapist_records:
        therapist_records.sort(key=lambda r: r.get('EffectiveDate', date.min), reverse=True)
        
        for record in therapist_records:
            effective_date = record.get('EffectiveDate')
            if effective_date and effective_date <= target_date:
                return record.get('Competency')
    
    # Fall back to current competency from Config_Therapists
    for therapist in config.get('therapists', []):
        if therapist.get('Name', '').strip().lower() == therapist_name.strip().lower():
            return therapist.get('Competency')
    
    return None


def get_competency_ranges_for_therapist(therapist_name, year, config):
    """
    Get column ranges for each competency period for a therapist.
    
    Returns list of tuples: [(competency, start_col, end_col), ...]
    
    Example for Chris (CA Jan, Senior from Feb):
    [('CA', 3, 3), ('Senior', 4, 15)]  # C for Jan, D-O for Feb-Dec+Avg
    """
    logging.info(f"  [DEBUG] get_competency_ranges called: name='{therapist_name}', year={year}")
    
    history = config.get('competency_history', [])
    logging.info(f"  [DEBUG] Total competency_history records: {len(history)}")
    
    therapist_records = [
        r for r in history
        if r.get('Name', '').strip().lower() == therapist_name.strip().lower()
        and r.get('EffectiveDate')
    ]
    
    logging.info(f"  [DEBUG] Records matching '{therapist_name}': {len(therapist_records)}")
    for r in therapist_records:
        logging.info(f"  [DEBUG]   Record: Name='{r.get('Name')}', Comp='{r.get('Competency')}', Date={r.get('EffectiveDate')} (type={type(r.get('EffectiveDate')).__name__})")
    
    if not therapist_records or not year:
        logging.info(f"  [DEBUG] Early return None: records={len(therapist_records) if therapist_records else 0}, year={year}")
        return None
    
    therapist_records.sort(key=lambda r: r.get('EffectiveDate'))
    
    # Determine competency for each month column (3-14 = Jan-Dec)
    month_competencies = {}
    for col in range(3, 15):  # C to N (Jan to Dec)
        month_name = COL_TO_MONTH.get(col)
        if not month_name:
            logging.info(f"  [DEBUG] col {col}: no month_name in COL_TO_MONTH")
            continue
        
        month_num = MONTH_NAME_TO_NUM.get(month_name)
        if not month_num:
            logging.info(f"  [DEBUG] col {col}, month '{month_name}': no month_num in MONTH_NAME_TO_NUM")
            continue
        
        target_date = date(year, month_num, 15)
        
        applicable_comp = None
        for record in therapist_records:
            eff_date = record.get('EffectiveDate')
            if hasattr(eff_date, 'date'):
                eff_date = eff_date.date()
            if eff_date <= target_date:
                applicable_comp = record.get('Competency')
        
        if applicable_comp:
            month_competencies[col] = applicable_comp
    
    logging.info(f"  [DEBUG] month_competencies: {month_competencies}")
    
    if not month_competencies:
        logging.info(f"  [DEBUG] No month_competencies, returning None")
        return None
    
    # Group consecutive columns with same competency into ranges
    ranges = []
    current_comp = None
    range_start_col = None
    range_end_col = None
    
    for col in range(3, 15):  # C to N
        if col not in month_competencies:
            continue
        comp = month_competencies[col]
        
        if comp == current_comp:
            range_end_col = col
        else:
            if current_comp is not None:
                ranges.append((current_comp, range_start_col, range_end_col))
            current_comp = comp
            range_start_col = col
            range_end_col = col
    
    # Don't forget the last range
    if current_comp is not None:
        ranges.append((current_comp, range_start_col, range_end_col))
    
    # Add Average column (15) to the last competency's range
    if ranges:
        last_comp, last_start, last_end = ranges[-1]
        ranges[-1] = (last_comp, last_start, 15)  # Extend to column O (Average)
    
    logging.info(f"  [DEBUG] Final ranges: {ranges}, len={len(ranges)}")
    result = ranges if len(ranges) > 1 else None
    logging.info(f"  [DEBUG] Returning: {result}")
    
    return result


def apply_billing_formatting_with_history(ws, therapist_name, year, config, thresholds_all, colours, kpi_row=4):
    """
    Apply billing formatting with competency history support.
    
    If therapist has competency changes during the year, applies different
    thresholds to different column ranges.
    
    Args:
        ws: Worksheet
        therapist_name: Name of therapist
        year: Year for competency lookup
        config: Full config dict
        thresholds_all: All thresholds for team type (e.g., config['thresholds']['Physio'])
        colours: Color config dict
        kpi_row: Row number for BillingsKPI (default 4)
    """
    # Check for competency history
    comp_ranges = get_competency_ranges_for_therapist(therapist_name, year, config)
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = colours.get('blue', DEFAULT_COLORS['blue'])
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    if comp_ranges:
        # Apply different thresholds to different column ranges
        logging.info(f"  {therapist_name}: {len(comp_ranges)} competency ranges")
        
        for competency, start_col, end_col in comp_ranges:
            thresholds = thresholds_all.get(competency, {})
            if not thresholds:
                logging.warning(f"No thresholds for competency {competency}")
                continue
            
            green_min = thresholds.get('green_min', 0)
            blue_above = thresholds.get('blue_above', 100)
            
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            cell_range = f'{start_letter}{kpi_row}:{end_letter}{kpi_row}'
            
            # Apply formatting rules
            ws.conditional_formatting.add(cell_range,
                FormulaRule(formula=[f'=LEN(TRIM({start_letter}{kpi_row}))=0'], fill=white_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=[str(blue_above)], fill=blue_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThanOrEqual', formula=[str(green_min)], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan', formula=[str(green_min)], fill=red_fill))
            
            logging.info(f"    {competency}: cols {start_letter}-{end_letter}, green>={green_min}, blue>{blue_above}")
    else:
        # No history - use standard single-range formatting with current competency
        # Get current competency from therapists list
        current_comp = None
        for t in config.get('therapists', []):
            if t.get('Name', '').strip().lower() == therapist_name.strip().lower():
                current_comp = t.get('Competency', 'CA')
                break
        
        thresholds = thresholds_all.get(current_comp, {})
        if not thresholds:
            logging.warning(f"No thresholds for {therapist_name} competency {current_comp}")
            return
        
        green_min = thresholds.get('green_min', 0)
        blue_above = thresholds.get('blue_above', 100)
        
        logging.info(f"Billing thresholds: green>={green_min}, blue>{blue_above}")
        
        billing_range = f'C{kpi_row}:O{kpi_row}'
        
        ws.conditional_formatting.add(billing_range,
            FormulaRule(formula=[f'=LEN(TRIM(C{kpi_row}))=0'], fill=white_fill))
        ws.conditional_formatting.add(billing_range,
            CellIsRule(operator='greaterThan', formula=[str(blue_above)], fill=blue_fill))
        ws.conditional_formatting.add(billing_range,
            CellIsRule(operator='greaterThanOrEqual', formula=[str(green_min)], fill=green_fill))
        ws.conditional_formatting.add(billing_range,
            CellIsRule(operator='lessThan', formula=[str(green_min)], fill=red_fill))


# =============================================================================
# TEMPLATE CREATION - NEW LAYOUT
# =============================================================================

def create_from_template_v2(therapist, config, token,
                            resolve_file_path, download_excel_file, upload_excel_file):
    """
    Create new therapist file from template - NEW LAYOUT.
    
    Writes:
    - Name to B2
    - Competency to D2
    
    Args:
        therapist: Therapist dict from config
        config: Full config dict (can be empty {})
        token: Graph API token
        resolve_file_path: Function to resolve SharePoint paths
        download_excel_file: Function to download Excel files
        upload_excel_file: Function to upload Excel files
        
    Returns:
        bool: Success/failure
    """
    name = therapist['Name']
    team_id = therapist['Team']
    team_type = 'OT' if team_id == 'OT' else 'Physio'
    competency = therapist.get('Competency', 'CA')
    file_path = therapist['FilePath']
    
    logging.info(f"Creating new file for {name} from {team_type} template (v2)")
    
    # Template path
    template_path = f"/Excel files/KPI Files/KPI/templates/Template_{team_type}.xlsx"
    
    # Download template
    template_file_id = resolve_file_path(template_path, token)
    if not template_file_id:
        logging.error(f"Template not found: {template_path}")
        return False
    
    try:
        template_content = download_excel_file(template_file_id, token)
        wb = load_workbook(template_content)
        
        # Find Dashboard sheet
        dashboard_sheet = None
        for sheet_name in wb.sheetnames:
            if 'Dashboard' in sheet_name and 'KPI' not in sheet_name:
                dashboard_sheet = sheet_name
                break
        
        if not dashboard_sheet:
            logging.error("Dashboard sheet not found in template")
            wb.close()
            return False
        
        ws = wb[dashboard_sheet]
        
        # Write therapist name to B2
        ws.cell(row=2, column=2, value=name)
        
        # Write competency to D2 (NEW)
        ws.cell(row=2, column=4, value=competency)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        wb.close()
        
        # Upload as new file
        upload_excel_file(file_path, output, token)
        logging.info(f"Created new file for {name} at {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to create file from template: {e}")
        return False


# =============================================================================
# CONDITIONAL FORMATTING FUNCTIONS - ROW-BASED
# =============================================================================

def apply_billing_formatting_v2(ws, thresholds, colours, kpi_row=4):
    """
    Apply conditional formatting to billing KPI row.
    
    NEW: Applies to a single row (C4:O4) instead of a column.
    
    Args:
        ws: Worksheet
        thresholds: Dict with red_below, green_min, green_max, blue_above
        colours: Color config dict
        kpi_row: Row number for BillingsKPI (default 4)
    """
    if not thresholds:
        logging.warning("No thresholds provided for billing formatting")
        return
    
    # Get threshold values from config
    green_min = thresholds.get('green_min', 0)
    blue_above = thresholds.get('blue_above', 100)
    
    logging.info(f"Billing thresholds: green>={green_min}, blue>{blue_above}")
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = colours.get('blue', DEFAULT_COLORS['blue'])
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    # Billing row range: C{row}:O{row} (includes Average column)
    billing_range = f'C{kpi_row}:O{kpi_row}'
    
    # 1. White for blank cells
    ws.conditional_formatting.add(billing_range,
        FormulaRule(formula=[f'=LEN(TRIM(C{kpi_row}))=0'], fill=white_fill))
    
    # 2. Blue for > blue_above (excellent)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='greaterThan', formula=[str(blue_above)], fill=blue_fill))
    
    # 3. Green for >= green_min (meets threshold)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(green_min)], fill=green_fill))
    
    # 4. Red for < green_min (below threshold)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='lessThan', formula=[str(green_min)], fill=red_fill))


def apply_ceased_services_formatting_v2(ws, ceased_thresholds, colours, kpi_row=5):
    """
    Apply conditional formatting to Ceased Services row (Physio only).
    
    NEW: Applies to a single row (C5:O5) instead of a column.
    Lower is better for ceased services.
    
    Args:
        ws: Worksheet
        ceased_thresholds: Dict with blue_below, green_min, green_max, red_above
        colours: Color config dict
        kpi_row: Row number for Ceased Services (default 5)
    """
    # Get threshold values from config
    blue_below = ceased_thresholds.get('blue_below', 0.025)
    red_above = ceased_thresholds.get('red_above', 0.04)
    
    logging.info(f"Ceased thresholds: blue<{blue_below}, red>={red_above}")
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = colours.get('blue', DEFAULT_COLORS['blue'])
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    # Ceased row range: C{row}:O{row} (includes Average column)
    ceased_range = f'C{kpi_row}:O{kpi_row}'
    
    # 1. White for blank
    ws.conditional_formatting.add(ceased_range,
        FormulaRule(formula=[f'=LEN(TRIM(C{kpi_row}))=0'], fill=white_fill))
    
    # 2. Blue for < blue_below (excellent - very low ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='lessThan', formula=[str(blue_below)], fill=blue_fill))
    
    # 3. Green for < red_above (good - acceptable ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='lessThan', formula=[str(red_above)], fill=green_fill))
    
    # 4. Red for >= red_above (poor - high ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(red_above)], fill=red_fill))


def apply_rating_scale_formatting_v2(ws, kpi_rows, rating_thresholds, colours):
    """
    Apply conditional formatting using 1-5 rating scale to KPI rows.
    
    NEW: Applies to row ranges (C{row}:O{row}) instead of column ranges.
    
    Args:
        ws: Worksheet
        kpi_rows: List of row numbers to format (e.g., [6, 7, 8])
        rating_thresholds: List of threshold dicts with rating, min, max
        colours: Color config dict
    """
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    amber_color = colours.get('amber', DEFAULT_COLORS['amber'])
    yellow_color = colours.get('yellow', DEFAULT_COLORS['yellow'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = colours.get('blue', DEFAULT_COLORS['blue'])
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    amber_fill = PatternFill(start_color=amber_color, end_color=amber_color, fill_type='solid')
    yellow_fill = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    color_map = {
        1: red_fill,
        2: amber_fill,
        3: yellow_fill,
        4: green_fill,
        5: blue_fill
    }
    
    for row in kpi_rows:
        cell_range = f'C{row}:O{row}'  # Includes Average column
        
        # 1. White for blank
        ws.conditional_formatting.add(cell_range,
            FormulaRule(formula=[f'=LEN(TRIM(C{row}))=0'], fill=white_fill))
        
        # Apply rules from config (highest rating first for priority)
        for threshold in reversed(rating_thresholds):
            rating = threshold['rating']
            min_val = threshold['min']
            max_val = threshold['max']
            fill = color_map.get(rating, white_fill)
            
            if rating == 5:
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='greaterThanOrEqual', formula=[str(min_val)], fill=fill))
            elif rating == 1:
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='lessThan', formula=[str(max_val)], fill=fill))
            else:
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='between', formula=[str(min_val), str(max_val)], fill=fill))


# =============================================================================
# MAIN FUNCTION - UPDATE INDIVIDUAL SHEET (NEW LAYOUT)
# =============================================================================

def update_individual_sheet_v2(therapist, config, master_data, token, 
                                resolve_file_path, download_excel_file, 
                                upload_excel_file, create_from_template, year=None):
    """
    Update a single therapist's individual KPI sheet - NEW TEMPLATE LAYOUT.
    
    Key difference from v1: Months are columns (C-N), KPIs are rows (4-8).
    
    Args:
        therapist: Therapist dict from config
        config: Full config dict
        master_data: KPI data from Team Leader file
        token: Graph API token
        resolve_file_path: Function to resolve SharePoint paths
        download_excel_file: Function to download Excel files
        upload_excel_file: Function to upload Excel files
        create_from_template: Function to create file from template
        year: Year for competency history lookup (e.g., 2026)
        
    Returns:
        bool: Success/failure
    """
    name = therapist['Name']
    team_id = therapist['Team']
    team_type = 'OT' if team_id == 'OT' else 'Physio'
    competency = therapist.get('Competency', 'CA')
    file_path = therapist['FilePath']
    
    logging.info(f"Updating {name} (v2 - new layout)...")
    
    # Resolve file path
    file_id = resolve_file_path(file_path, token)
    
    if not file_id:
        logging.warning(f"File not found: {file_path} - attempting to create from template")
        if create_from_template(therapist, config, token):
            file_id = resolve_file_path(file_path, token)
            if not file_id:
                logging.error(f"File still not found after template creation: {file_path}")
                return False
        else:
            logging.error(f"Failed to create file from template: {file_path}")
            return False
    
    # Get KPI data for this therapist
    team_data = master_data.get(team_id, [])
    kpi_records = [r for r in team_data if r.get('Name') == name or r.get('UniqueTeamMembers') == name]
    
    if not kpi_records:
        logging.warning(f"No KPI data for {name} in team {team_id} - continuing to update thresholds")
    
    # Download file
    file_content = download_excel_file(file_id, token)
    wb = load_workbook(file_content)
    
    # Find Dashboard sheet
    dashboard_sheet = None
    for sheet_name in wb.sheetnames:
        if 'Dashboard' in sheet_name and 'KPI' not in sheet_name:
            dashboard_sheet = sheet_name
            break
    
    if not dashboard_sheet:
        logging.error(f"Dashboard sheet not found in {file_path}")
        wb.close()
        return False
    
    ws = wb[dashboard_sheet]
    
    # Select KPI row mapping based on team type
    kpi_rows = KPI_ROWS_OT if team_type == 'OT' else KPI_ROWS_PHYSIO
    
    # =========================================================================
    # WRITE KPI DATA (months as columns, KPIs as rows)
    # =========================================================================
    for record in kpi_records:
        month = record.get('Month')
        if month not in MONTH_COLUMNS:
            continue
        
        col = MONTH_COLUMNS[month]
        
        for kpi_name, row in kpi_rows.items():
            value = record.get(kpi_name)
            if value is not None:
                ws.cell(row=row, column=col, value=value)
    
    # =========================================================================
    # APPLY CELL FORMATTING (number format, alignment, font)
    # =========================================================================
    
    # Define standard formatting
    center_align = Alignment(horizontal='center', vertical='center')
    black_font = Font(color='000000')
    
    # Format all KPI data cells (rows 4-8, columns C-O)
    for row in range(4, 9):
        for col in range(3, 16):  # C to O
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.font = black_font
    
    # BillingsKPI row (row 4): 2 decimal places for average
    billings_avg_cell = ws.cell(row=4, column=15)  # O4
    billings_avg_cell.number_format = '0.00'
    
    # Ceased Services row (row 5, Physio only): percentage format
    if team_type == 'Physio':
        for col in range(3, 16):  # C to O
            cell = ws.cell(row=5, column=col)
            cell.number_format = '0.0%'
        
        # Set correct average formula for Ceased Services (include 0s, exclude blanks)
        ceased_avg_cell = ws.cell(row=5, column=15)  # O5
        ceased_avg_cell.value = '=IFERROR(AVERAGEIF($C5:$N5,"<>"),"")' 
        
        # Rating scale KPI averages (rows 6-8): 2 decimal places
        for row in [6, 7, 8]:
            ws.cell(row=row, column=15).number_format = '0.00'
    else:
        # OT: Rating scale KPI averages (rows 5-8): 2 decimal places
        for row in [5, 6, 7, 8]:
            ws.cell(row=row, column=15).number_format = '0.00'
    
    # =========================================================================
    # WRITE THERAPIST INFO
    # =========================================================================
    # Name in B2
    ws.cell(row=2, column=2, value=name)
    
    # Competency in D2
    ws.cell(row=2, column=4, value=competency)
    
    # =========================================================================
    # WRITE THRESHOLD DISPLAY VALUES
    # =========================================================================
    thresholds_all = config['thresholds'].get(team_type, {})
    
    # Billing thresholds: Rows 12-14, Columns C-E
    for comp_level, row in THRESHOLD_ROWS['billings'].items():
        comp_thresholds = thresholds_all.get(comp_level, {})
        if comp_thresholds:
            try:
                # Column C: Below threshold
                ws.cell(row=row, column=3, value=f"<{comp_thresholds.get('green_min', '')}")
                # Column D: Good range
                ws.cell(row=row, column=4, value=f"{comp_thresholds.get('green_min', '')}-{comp_thresholds.get('green_max', '')}")
                # Column E: Excellent
                ws.cell(row=row, column=5, value=f">{comp_thresholds.get('blue_above', '')}")
            except Exception as e:
                logging.warning(f"Could not write {comp_level} thresholds: {e}")
    
    # Ceased Services thresholds: Rows 17-19, Column E (Physio only)
    # Note: Column D has labels (Excellent/Good/Below), values go in E
    if team_type == 'Physio':
        ceased = config.get('ceased_thresholds', {})
        blue_below = ceased.get('blue_below', 0.025)
        red_above = ceased.get('red_above', 0.04)
        
        try:
            # Row 17: Excellent (< blue_below)
            ws.cell(row=17, column=2, value=f"<{blue_below*100:.1f}%")
            # Row 18: Good (between)
            ws.cell(row=18, column=2, value=f"{blue_below*100:.1f}-{red_above*100:.1f}%")
            # Row 19: Below (>= red_above)
            ws.cell(row=19, column=2, value=f">{red_above*100:.1f}%")
        except Exception as e:
            logging.warning(f"Could not write ceased thresholds: {e}")
    
    # =========================================================================
    # APPLY CONDITIONAL FORMATTING
    # =========================================================================
    
    # Clear existing conditional formatting
    ws.conditional_formatting._cf_rules = {}
    
    if team_type == 'OT':
        # OT: Billings (row 4) with competency history, then rating scale (rows 5-8)
        apply_billing_formatting_with_history(ws, name, year, config, thresholds_all, config['colours'], kpi_row=4)
        apply_rating_scale_formatting_v2(ws, [5, 6, 7, 8], config['rating_thresholds'], config['colours'])
    else:
        # Physio: Billings (row 4) with competency history, Ceased (row 5), then rating scale (rows 6-8)
        apply_billing_formatting_with_history(ws, name, year, config, thresholds_all, config['colours'], kpi_row=4)
        apply_ceased_services_formatting_v2(ws, config['ceased_thresholds'], config['colours'], kpi_row=5)
        apply_rating_scale_formatting_v2(ws, [6, 7, 8], config['rating_thresholds'], config['colours'])
    
    # =========================================================================
    # SAVE AND UPLOAD
    # =========================================================================
    output = BytesIO()
    wb.save(output)
    wb.close()
    
    upload_excel_file(file_path, output, token)
    logging.info(f"Updated {name} (v2)")
    return True


# =============================================================================
# WRAPPER FOR EASY INTEGRATION
# =============================================================================

def create_from_template(therapist, config, token):
    """
    Wrapper that imports dependencies and calls create_from_template_v2.
    
    Drop-in replacement for the original function.
    Works with both function_app.py (Azure) and function_app_local.py (local testing).
    """
    try:
        from function_app import (
            resolve_file_path,
            download_excel_file,
            upload_excel_file
        )
    except ImportError:
        from function_app_local import (
            resolve_file_path,
            download_excel_file,
            upload_excel_file
        )
    
    return create_from_template_v2(
        therapist, config, token,
        resolve_file_path, download_excel_file, upload_excel_file
    )


def update_individual_sheet(therapist, config, master_data, token, year=None):
    """
    Wrapper that imports dependencies and calls update_individual_sheet_v2.
    
    Drop-in replacement for the original function.
    Works with both function_app.py (Azure) and function_app_local.py (local testing).
    
    Args:
        therapist: Therapist dict from config
        config: Full config dict
        master_data: KPI data from Team Leader file
        token: Graph API token
        year: Year for competency history lookup (e.g., 2026)
    """
    try:
        from function_app import (
            resolve_file_path,
            download_excel_file,
            upload_excel_file
        )
    except ImportError:
        from function_app_local import (
            resolve_file_path,
            download_excel_file,
            upload_excel_file
        )
    
    # Use local create_from_template which calls v2
    return update_individual_sheet_v2(
        therapist, config, master_data, token,
        resolve_file_path, download_excel_file,
        upload_excel_file, create_from_template, year
    )
