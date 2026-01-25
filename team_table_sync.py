"""
Team Table Sync Module
Dynamically add/remove/reorder therapist rows in Team Leader KPI tables
based on Config_Therapists.

Features:
- Add new therapists to all KPI tables for their team
- Remove therapists no longer in config
- Remove placeholder rows
- Grey out inactive therapists (IsActive = FALSE)
- Order: Team Leaders first, then by Competency (Senior Ã¢â€ â€™ CA Ã¢â€ â€™ Grad)

Row change handling:
- If rows need DELETING: clears values only, preserves table structure for manual deletion
- If rows need ADDING: skips sync for that team, notifies admin
- Emails only sent on Monday/Thursday around 10 AM
"""

import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.utils import range_boundaries, get_column_letter
from copy import copy


# =============================================================================
# CONFIGURATION
# =============================================================================

# Competency sort order (lower = first)
COMPETENCY_ORDER = {
    'Senior': 1,
    'CA': 2,
    'Grad': 3
}

# Grey fill for inactive therapists
INACTIVE_FILL = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')
INACTIVE_FONT = Font(color='FF808080')

# Team to sheet mapping
TEAM_SHEET_MAP = {
    'Physio_North': 'KPI Dashboard North',
    'Physio_South': 'KPI Dashboard South',
    'OT': 'KPI Dashboard OT'
}

# Tables per team (in order they appear on sheet)
TEAM_TABLES = {
    'Physio_North': ['Billings_North', 'Ceased_North', 'Documentation_North', 'Admin_North', 'Attitude_North'],
    'Physio_South': ['Billings_South', 'Ceased_South', 'Documentation_South', 'Admin_South', 'Attitude_South'],
    'OT': ['Billings_OT', 'Compliance_OT', 'ReferrerEng_OT', 'Capacity_OT', 'Attitude_OT']
}

# Email notification settings
NOTIFICATION_EMAIL = "isaac@melbournemobilephysio.net.au"  # TODO: Update this
EMAIL_DAYS = [0, 3]  # Monday=0, Thursday=3
EMAIL_HOUR_MIN = 9   # 9 AM
EMAIL_HOUR_MAX = 12  # 11 AM


# =============================================================================
# EMAIL NOTIFICATION HELPERS
# =============================================================================

def is_email_window():
    """Check if current time is within the email notification window.
    
    Returns True if:
    - Day is Monday (0) or Thursday (3)
    - Hour is between 9 AM and 11 AM (inclusive)
    """
    now = datetime.now()
    day_of_week = now.weekday()  # Monday=0, Sunday=6
    hour = now.hour
    
    if day_of_week in EMAIL_DAYS and EMAIL_HOUR_MIN <= hour <= EMAIL_HOUR_MAX:
        logging.info(f"Within email window: {now.strftime('%A %H:%M')}")
        return True
    
    logging.debug(f"Outside email window: {now.strftime('%A %H:%M')}")
    return False


def send_email_graph_api(token, subject, body_html):
    """
    Send email using Microsoft Graph API.
    
    Requires Mail.Send application permission on the app registration.
    
    Args:
        token: Graph API access token
        subject: Email subject line
        body_html: Email body (HTML format)
        
    Returns:
        bool: True if sent successfully
    """
    import requests
    
    # Graph API endpoint - send as the NOTIFICATION_EMAIL user
    url = f"https://graph.microsoft.com/v1.0/users/{NOTIFICATION_EMAIL}/sendMail"
    
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    email_data = {
        "message": {
            "subject": subject,
            "body": {
                "contentType": "HTML",
                "content": body_html
            },
            "toRecipients": [
                {
                    "emailAddress": {
                        "address": NOTIFICATION_EMAIL
                    }
                }
            ]
        },
        "saveToSentItems": "true"
    }
    
    try:
        response = requests.post(url, headers=headers, json=email_data)
        
        if response.status_code == 202:
            logging.info(f"Email sent successfully to {NOTIFICATION_EMAIL}")
            return True
        else:
            logging.error(f"Failed to send email: {response.status_code} - {response.text}")
            return False
            
    except Exception as e:
        logging.error(f"Error sending email: {e}")
        return False


def send_consolidated_email(pending_changes, token=None):
    """
    Send a single consolidated email for all row changes needed.
    
    Args:
        pending_changes: List of dicts with keys:
            - sheet_name: e.g., 'KPI Dashboard North'
            - team_name: e.g., 'Physio_North'
            - action: 'add' or 'delete'
            - row_count: number of rows
        token: Graph API access token (optional)
        
    Returns:
        bool: True if sent successfully
    """
    if not pending_changes:
        return False
    
    if not is_email_window():
        logging.info("Row changes needed but outside email window - no email sent")
        return False
    
    subject = "KPI Tables: Manual action required"
    
    # Build HTML body
    sections = []
    
    for change in pending_changes:
        sheet_name = change['sheet_name']
        team_name = change['team_name']
        action = change['action']
        row_count = change['row_count']
        
        if action == 'add':
            action_text = '<b>added</b>'
            steps = f"""<p><b>Steps:</b></p>
<ol>
    <li>Open the Team Leader file in Excel</li>
    <li>Go to sheet: {sheet_name}</li>
    <li>For each table, right-click in the table and select "Insert Table Rows"</li>
    <li>Add {row_count} new row(s)</li>
</ol>
<p>Once rows are added, the next sync will populate the data automatically.</p>"""
        else:  # delete
            action_text = '<b>deleted</b>'
            steps = f"""<p><b>Steps:</b></p>
<ol>
    <li>Open the Team Leader file in Excel</li>
    <li>Go to sheet: {sheet_name}</li>
    <li>For each table, right-click on the empty rows and select "Delete Table Rows"</li>
    <li>Delete {row_count} blank row(s)</li>
</ol>"""
        
        section = f"""<div style="margin-bottom: 20px; padding: 10px; border-left: 3px solid #0078d4;">
    <p><b>Sheet:</b> {sheet_name}<br>
    <b>Team:</b> {team_name}</p>
    <p>{row_count} row(s) need to be {action_text} from the team tables.</p>
    {steps}
</div>"""
        sections.append(section)
    
    body_html = f"""<html>
<body style="font-family: Arial, sans-serif;">
<p>Manual action required for Team Leader KPI Dashboard.</p>

{''.join(sections)}

<hr>
<p style="color: #666; font-size: 12px;">This is an automated message from the KPI Processing System.</p>
</body>
</html>"""
    
    if token:
        return send_email_graph_api(token, subject, body_html)
    else:
        logging.warning(f"EMAIL (no token provided) - Would send:")
        logging.warning(f"  Subject: {subject}")
        logging.warning(f"  Changes: {len(pending_changes)} teams affected")
        return False


# =============================================================================
# HELPER: GET SORTED THERAPISTS FOR TEAM
# =============================================================================

def get_therapists_for_team(config, team_name):
    """
    Get therapists for a team, sorted by: Team Leaders first, then Competency.
    
    Args:
        config: Config dict with 'therapists' list
        team_name: 'Physio_North' | 'Physio_South' | 'OT'
        
    Returns:
        list of therapist dicts, sorted appropriately
    """
    team_therapists = [
        t for t in config.get('therapists', [])
        if t.get('Team') == team_name
    ]
    
    # Sort: Team Leaders first (True before False), then by Competency order
    def sort_key(t):
        is_leader = t.get('IsTeamLeader', False)
        # Convert to sortable: True = 0 (first), False = 1 (second)
        leader_sort = 0 if is_leader else 1
        competency_sort = COMPETENCY_ORDER.get(t.get('Competency', 'Grad'), 99)
        name_sort = t.get('Name', '')
        return (leader_sort, competency_sort, name_sort)
    
    sorted_therapists = sorted(team_therapists, key=sort_key)
    
    logging.info(f"Team {team_name}: {len(sorted_therapists)} therapists")
    for t in sorted_therapists:
        leader_flag = " [LEADER]" if t.get('IsTeamLeader') else ""
        active_flag = "" if str(t.get('IsActive', 'TRUE')).upper() == 'TRUE' else " [INACTIVE]"
        logging.debug(f"  {t['Name']} ({t.get('Competency')}){leader_flag}{active_flag}")
    
    return sorted_therapists


# =============================================================================
# HELPER: GET CURRENT TABLE ROWS
# =============================================================================

def get_table_rows(ws, table_name):
    """
    Get current row data from a table.
    
    Returns:
        list of dicts: [{'row_idx': 14, 'name': 'Chris', 'data': [values...]}, ...]
    """
    if table_name not in ws.tables:
        logging.warning(f"Table '{table_name}' not found")
        return []
    
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    rows = []
    for row_idx in range(min_row + 1, max_row + 1):  # Skip header
        name = ws.cell(row=row_idx, column=min_col).value
        if name:
            # Get all cell values in the row
            data = []
            for col_idx in range(min_col, max_col + 1):
                data.append(ws.cell(row=row_idx, column=col_idx).value)
            rows.append({
                'row_idx': row_idx,
                'name': str(name).strip(),
                'data': data
            })
    
    return rows


# =============================================================================
# HELPER: CHECK IF ROW IS PLACEHOLDER
# =============================================================================

def is_placeholder_row(name):
    """Check if a row name indicates it's a placeholder."""
    if not name:
        return True
    name_upper = str(name).upper()
    return 'PLACEHOLDER' in name_upper or name_upper.startswith('PLACEHOLDER')



# =============================================================================
# SYNC ALL TABLES FOR A TEAM
# =============================================================================

def sync_team_tables(ws, team_name, config):
    """
    Sync all KPI tables for a team.
    
    Handles three scenarios:
    1. Rows need ADDING: Skip sync, return pending change
    2. Rows need DELETING: Clear values only (preserve table structure), return pending change
    3. Same row count: Reorder/update data normally
    
    Args:
        ws: Worksheet object
        team_name: 'Physio_North' | 'Physio_South' | 'OT'
        config: Full config dict
        
    Returns:
        dict: Statistics including 'pending_change' if manual action needed
    """
    logging.info(f"Syncing tables for {team_name}")
    sheet_name = TEAM_SHEET_MAP.get(team_name, team_name)
    
    # Get sorted therapists for this team
    therapists = get_therapists_for_team(config, team_name)
    
    if not therapists:
        logging.warning(f"No therapists found for team {team_name}")
        return {'tables': 0, 'added': 0, 'removed': 0, 'skipped': False}
    
    # Get tables for this team
    table_names = TEAM_TABLES.get(team_name, [])
    new_row_count = len(therapists)
    therapist_info = {t['Name']: t for t in therapists}
    
    # Check first table to determine current row count
    first_table_name = table_names[0] if table_names else None
    if not first_table_name or first_table_name not in ws.tables:
        logging.warning(f"First table {first_table_name} not found")
        return {'tables': 0, 'added': 0, 'removed': 0, 'skipped': True}
    
    first_table = ws.tables[first_table_name]
    min_col, min_row, max_col, max_row = range_boundaries(first_table.ref)
    current_row_count = max_row - min_row  # Excludes header
    
    row_delta = new_row_count - current_row_count
    
    logging.info(f"  Current rows: {current_row_count}, Expected: {new_row_count}, Delta: {row_delta}")
    
    # ==========================================================================
    # SCENARIO 1: Need to ADD rows - cannot do automatically, notify and skip
    # ==========================================================================
    if row_delta > 0:
        logging.warning(f"  {team_name}: Need to add {row_delta} rows - SKIPPING SYNC")
        return {
            'tables': 0,
            'added': 0,
            'removed': 0,
            'skipped': True,
            'reason': f'Need {row_delta} more rows',
            'pending_change': {
                'sheet_name': sheet_name,
                'team_name': team_name,
                'action': 'add',
                'row_count': row_delta
            }
        }
    
    # ==========================================================================
    # SCENARIO 2 & 3: Same rows or need to DELETE rows - process tables
    # ==========================================================================
    
    total_stats = {'tables': 0, 'added': 0, 'removed': 0, 'skipped': False}
    rows_cleared = 0
    
    for table_name in table_names:
        if table_name not in ws.tables:
            logging.warning(f"Table {table_name} not found")
            continue
        
        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        col_count = max_col - min_col + 1
        
        # Get current data
        current_rows = get_table_rows(ws, table_name)
        current_data = {r['name']: r['data'] for r in current_rows}
        old_row_count = len(current_rows)
        
        # Count changes
        current_names = set(current_data.keys())
        expected_names = {t['Name'] for t in therapists}
        placeholders = {n for n in current_names if is_placeholder_row(n)}
        
        added = len(expected_names - current_names)
        removed = len((current_names - expected_names) | placeholders)
        
        total_stats['added'] += added
        total_stats['removed'] += removed
        total_stats['tables'] += 1
        
        # NOTE: Do NOT shrink table reference - leave rows for manual deletion
        # This preserves table structure and formatting
        
        # Write therapist data in correct order
        first_data_row = min_row + 1
        for i, therapist in enumerate(therapists):
            row_idx = first_data_row + i
            name = therapist['Name']
            is_active = therapist.get('IsActive', True)
            
            # Get existing data or create empty
            row_data = current_data.get(name, [name] + [None] * (col_count - 1))
            
            # Skip last column (col_count - 1) to preserve average formulas
            for col_offset in range(col_count - 1):
                cell = ws.cell(row=row_idx, column=min_col + col_offset)
                
                if col_offset == 0:
                    cell.value = name
                elif name in current_data and col_offset < len(row_data):
                    cell.value = row_data[col_offset]
                else:
                    cell.value = None
                
                # Apply styling based on active status
                if not is_active:
                    cell.fill = INACTIVE_FILL
                    cell.font = INACTIVE_FONT
                else:
                    cell.fill = PatternFill()  # Clear any static fill
                    cell.font = Font()  # Reset font to default
        
        # Clear VALUES only in extra rows (preserve table structure for manual deletion)
        # Skip last column (max_col) to preserve average formulas
        if row_delta < 0:
            rows_to_clear = abs(row_delta)
            for row_idx in range(first_data_row + new_row_count, max_row + 1):
                for col_idx in range(min_col, max_col):  # Exclude max_col (average column)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None
                    # Keep borders and table formatting intact
            rows_cleared = rows_to_clear
        
        logging.info(f"  {table_name}: {old_row_count} -> {new_row_count} rows (cleared {rows_cleared if row_delta < 0 else 0})")
    
    # Return pending change if rows were cleared (need manual deletion)
    if rows_cleared > 0:
        logging.warning(f"  {team_name}: Cleared {rows_cleared} rows - need manual deletion")
        total_stats['pending_change'] = {
            'sheet_name': sheet_name,
            'team_name': team_name,
            'action': 'delete',
            'row_count': rows_cleared
        }
    
    logging.info(f"  Completed {team_name}: {total_stats['tables']} tables synced")
    return total_stats


# =============================================================================
# FTE TABLE SYNC
# =============================================================================

def sync_fte_table(wb, config):
    """
    Sync FTETable on FTE sheet with therapist data from config.
    
    Writes columns A-C only (Therapist, FTE, Team).
    Columns D-O contain formulas and are NOT touched.
    
    Unlike team tables, FTETable can auto-resize (add/remove rows).
    
    Args:
        wb: Workbook object
        config: Full config dict
        
    Returns:
        dict: Statistics
    """
    logging.info("Syncing FTETable...")
    
    sheet_name = 'FTE'
    table_name = 'FTETable'
    
    if sheet_name not in wb.sheetnames:
        logging.warning(f"Sheet '{sheet_name}' not found - skipping FTE sync")
        return {'rows': 0, 'skipped': True}
    
    ws = wb[sheet_name]
    
    if table_name not in ws.tables:
        logging.warning(f"Table '{table_name}' not found - skipping FTE sync")
        return {'rows': 0, 'skipped': True}
    
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    # Get all active therapists, sorted by team then competency
    therapists = [t for t in config.get('therapists', []) if t.get('IsActive', True)]
    
    # Sort: by team, then team leaders first, then competency
    def sort_key(t):
        team_order = {'Physio_North': 1, 'Physio_South': 2, 'OT': 3}
        is_leader = 0 if t.get('IsTeamLeader', False) else 1
        comp_order = COMPETENCY_ORDER.get(t.get('Competency', 'Grad'), 99)
        return (team_order.get(t.get('Team', ''), 99), is_leader, comp_order, t.get('Name', ''))
    
    therapists.sort(key=sort_key)
    
    new_row_count = len(therapists)
    current_row_count = max_row - min_row  # Excludes header
    row_delta = new_row_count - current_row_count
    
    logging.info(f"  FTETable: {current_row_count} rows -> {new_row_count} rows (delta: {row_delta})")
    
    # Calculate new table boundaries
    new_max_row = min_row + new_row_count
    first_data_row = min_row + 1
    
    # Write therapist data to columns A, B, C
    for i, therapist in enumerate(therapists):
        row_idx = first_data_row + i
        
        # Column A (min_col): Therapist name
        ws.cell(row=row_idx, column=min_col).value = therapist.get('Name', '')
        
        # Column B (min_col + 1): FTE
        ws.cell(row=row_idx, column=min_col + 1).value = therapist.get('FTE', 1)
        
        # Column C (min_col + 2): Team
        ws.cell(row=row_idx, column=min_col + 2).value = therapist.get('Team', '')
    
    # Clear extra rows if shrinking (columns A-C only)
    rows_cleared = 0
    if row_delta < 0:
        rows_cleared = abs(row_delta)
        for row_idx in range(new_max_row + 1, max_row + 1):
            for col_idx in range(min_col, min_col + 3):
                ws.cell(row=row_idx, column=col_idx).value = None
    
    # Update table reference to new size
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    table.ref = new_ref
    logging.info(f"  FTETable: Updated table ref to {new_ref}")
    
    logging.info(f"  FTETable: Synced {new_row_count} therapists")
    
    return {
        'rows': new_row_count,
        'added': max(0, row_delta),
        'removed': rows_cleared,
        'skipped': False
    }


# =============================================================================
# MAIN: SYNC ALL TEAM SHEETS
# =============================================================================

def sync_all_team_tables(wb, config, token=None):
    """
    Sync all team tables across all sheets.
    
    Args:
        wb: Workbook object
        config: Full config dict
        token: Graph API token for sending email notifications (optional)
        
    Returns:
        dict: Overall statistics
    """
    logging.info("="*60)
    logging.info("Starting Team Table Sync")
    logging.info("="*60)
    
    total_stats = {
        'teams': 0,
        'tables': 0,
        'added': 0,
        'removed': 0
    }
    
    pending_changes = []
    
    # Sync team dashboard tables
    for team_name, sheet_name in TEAM_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            logging.warning(f"Sheet '{sheet_name}' not found - skipping {team_name}")
            continue
        
        ws = wb[sheet_name]
        stats = sync_team_tables(ws, team_name, config)
        
        total_stats['teams'] += 1
        total_stats['tables'] += stats['tables']
        total_stats['added'] += stats['added']
        total_stats['removed'] += stats['removed']
        
        # Collect pending changes for consolidated email
        if 'pending_change' in stats:
            pending_changes.append(stats['pending_change'])
    
    # Sync FTE table
    fte_stats = sync_fte_table(wb, config)
    if 'pending_change' in fte_stats:
        pending_changes.append(fte_stats['pending_change'])
    
    logging.info("="*60)
    logging.info(f"SYNC COMPLETE: {total_stats['teams']} teams, {total_stats['tables']} tables, "
                 f"+{total_stats['added']} added, -{total_stats['removed']} removed")
    logging.info("="*60)
    
    # Send consolidated email if there are pending changes
    if pending_changes:
        logging.info(f"Pending manual changes: {len(pending_changes)} team(s) affected")
        send_consolidated_email(pending_changes, token)
    
    return total_stats


# =============================================================================
# STANDALONE TEST
# =============================================================================

if __name__ == "__main__":
    print("="*60)
    print("Team Table Sync Module")
    print("="*60)
    print("\nUsage:")
    print("  from team_table_sync import sync_all_team_tables")
    print("  stats = sync_all_team_tables(wb, config)")
    print("\nFeatures:")
    print("  - Adds new therapists from Config_Therapists")
    print("  - Removes therapists not in config")
    print("  - Removes placeholder rows")
    print("  - Grey out inactive therapists")
    print("  - Orders: Team Leaders first, then Senior Ã¢â€ â€™ CA Ã¢â€ â€™ Grad")
    print("\nTables synced per team:")
    for team, tables in TEAM_TABLES.items():
        print(f"  {team}: {', '.join(tables)}")
