"""
function_app_local.py
Local testing version - Azure Functions decorators removed
"""

# import azure.functions as func  # REMOVED for local testing
import logging
import json
import os
import re
from datetime import datetime, timedelta, date
from io import BytesIO
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# app = func.FunctionApp()  # REMOVED for local testing

# Cache for resolved file IDs
_file_id_cache = {}
_cache_expiry = None

# =============================================================================
# CONFIGURATION
# =============================================================================

SHAREPOINT_SITE_ID = os.environ.get('SHAREPOINT_SITE_ID', 'melbournemobilephysio.sharepoint.com,342fb7ab-2908-4049-b6bb-a2efa3cfdf8e,02f5c854-eb73-489e-890c-11d32dec607f')
DRIVE_ID = os.environ.get('DRIVE_ID') or os.environ.get('SHAREPOINT_DRIVE_ID', 'b!q7cvNAgpSUC2u6Lvo8_fjlTI9QJz655IiQwR0y3sYH82gtu7kzKKS4RzCsfD_vcW')

CONFIG_FILE_PATH = '/Admin -  Team Leaders/Team Performance and KPIs/MMP Team Performance Data.xlsx'

# Template file paths
TEMPLATE_PATHS = {
    'Physio': '/Admin -  Team Leaders/Templates/Template_Physio.xlsx',
    'OT': '/Admin -  Team Leaders/Templates/Template_OT.xlsx'
}

CONFIG_SHEETS = {
    'therapists': 'Config_Therapists',
    'teams': 'Config_Teams',
    'thresholds_physio': 'Config_Thresholds_Physio',
    'thresholds_ot': 'Config_Thresholds_OT',
    'kpis': 'Config_KPIs',
    'colours': 'Config_Colours',
    'competency_history': 'Config_Competency_History',
    'team_ave_thresholds': 'Config_TeamAve_Thresholds'  # Team average threshold history
}

# Default colours (will be overridden by config)
DEFAULT_COLORS = {
    'red': 'FFE47373',
    'amber': 'FFFFB74D', 
    'green': 'FF81C784',
    'grey': 'FFC0C0C0',
    'white': 'FFFFFFFF'
}

# =============================================================================
# GRAPH API HELPERS
# =============================================================================

def get_access_token():
    """Get access token using managed identity or app registration."""
    identity_endpoint = os.environ.get('IDENTITY_ENDPOINT')
    identity_header = os.environ.get('IDENTITY_HEADER')
    
    if identity_endpoint and identity_header:
        token_url = f"{identity_endpoint}?resource=https://graph.microsoft.com&api-version=2019-08-01"
        headers = {'X-IDENTITY-HEADER': identity_header}
        response = requests.get(token_url, headers=headers)
        if response.status_code == 200:
            return response.json()['access_token']
    
    tenant_id = os.environ.get('TENANT_ID') or os.environ.get('AZURE_TENANT_ID')
    client_id = os.environ.get('CLIENT_ID') or os.environ.get('AZURE_CLIENT_ID')
    client_secret = os.environ.get('CLIENT_SECRET') or os.environ.get('AZURE_CLIENT_SECRET')
    
    if tenant_id and client_id and client_secret:
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        data = {
            'grant_type': 'client_credentials',
            'client_id': client_id,
            'client_secret': client_secret,
            'scope': 'https://graph.microsoft.com/.default'
        }
        response = requests.post(token_url, data=data)
        if response.status_code == 200:
            return response.json()['access_token']
    
    raise Exception("Failed to obtain access token")


def graph_request(endpoint, token, method='GET', data=None, content_type='application/json'):
    """Make a request to Microsoft Graph API."""
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': content_type
    }
    
    url = f"https://graph.microsoft.com/v1.0{endpoint}"
    
    if method == 'GET':
        response = requests.get(url, headers=headers)
    elif method == 'PUT':
        response = requests.put(url, headers=headers, data=data)
    elif method == 'POST':
        response = requests.post(url, headers=headers, json=data if content_type == 'application/json' else None, data=data if content_type != 'application/json' else None)
    elif method == 'PATCH':
        response = requests.patch(url, headers=headers, json=data)
    
    return response


def resolve_file_path(file_path, token):
    """Resolve a SharePoint file path to a file ID, with caching."""
    global _file_id_cache, _cache_expiry
    
    now = datetime.utcnow()
    if _cache_expiry and now < _cache_expiry and file_path in _file_id_cache:
        return _file_id_cache[file_path]
    
    if not _cache_expiry or now >= _cache_expiry:
        _file_id_cache = {}
        _cache_expiry = now + timedelta(hours=24)
    
    encoded_path = file_path.replace(' ', '%20').replace('&', '%26')
    endpoint = f"/drives/{DRIVE_ID}/root:{encoded_path}"
    
    response = graph_request(endpoint, token)
    
    if response.status_code == 404:
        return None  # File doesn't exist
    
    if response.status_code >= 400:
        raise Exception(f"Graph API error: {response.status_code} - {response.text}")
    
    file_info = response.json()
    file_id = file_info['id']
    
    _file_id_cache[file_path] = file_id
    return file_id


def file_exists(file_path, token):
    """Check if a file exists at the given path."""
    file_id = resolve_file_path(file_path, token)
    return file_id is not None


def download_excel_file(file_id, token):
    """Download an Excel file from SharePoint."""
    endpoint = f"/drives/{DRIVE_ID}/items/{file_id}/content"
    response = graph_request(endpoint, token)
    if response.status_code >= 400:
        raise Exception(f"Failed to download file: {response.status_code}")
    return BytesIO(response.content)


def upload_excel_file(file_path, file_content, token):
    """Upload an Excel file to SharePoint (creates or overwrites)."""
    encoded_path = file_path.replace(' ', '%20').replace('&', '%26')
    endpoint = f"/drives/{DRIVE_ID}/root:{encoded_path}:/content"
    
    if isinstance(file_content, BytesIO):
        file_content.seek(0)
        data = file_content.read()
    else:
        data = file_content
    
    response = graph_request(endpoint, token, method='PUT', data=data,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    if response.status_code >= 400:
        raise Exception(f"Failed to upload file: {response.status_code} - {response.text}")
    
    return response


def create_folder(folder_path, token):
    """Create a folder in SharePoint if it doesn't exist."""
    encoded_path = folder_path.replace(' ', '%20').replace('&', '%26')
    endpoint = f"/drives/{DRIVE_ID}/root:{encoded_path}"
    response = graph_request(endpoint, token)
    
    if response.status_code == 200:
        return response.json()['id']
    
    parent_path = '/'.join(folder_path.rsplit('/', 1)[:-1]) or '/'
    folder_name = folder_path.rsplit('/', 1)[-1]
    
    encoded_parent = parent_path.replace(' ', '%20').replace('&', '%26')
    endpoint = f"/drives/{DRIVE_ID}/root:{encoded_parent}:/children"
    
    data = {
        'name': folder_name,
        'folder': {},
        '@microsoft.graph.conflictBehavior': 'fail'
    }
    
    response = graph_request(endpoint, token, method='POST', data=data)
    
    if response.status_code >= 400:
        raise Exception(f"Failed to create folder: {response.status_code}")
    
    return response.json()['id']


def move_file(file_id, new_parent_path, new_name, token):
    """Move a file to a new location with a new name."""
    parent_id = create_folder(new_parent_path, token)
    
    endpoint = f"/drives/{DRIVE_ID}/items/{file_id}"
    data = {
        'parentReference': {'id': parent_id},
        'name': new_name
    }
    
    response = graph_request(endpoint, token, method='PATCH', data=data)
    
    if response.status_code >= 400:
        raise Exception(f"Failed to move file: {response.status_code}")
    
    return response


# =============================================================================
# CONFIG LOADING
# =============================================================================

def load_config(token):
    """Load all configuration from the master Excel file."""
    logging.info("Loading configuration...")
    
    file_id = resolve_file_path(CONFIG_FILE_PATH, token)
    if not file_id:
        raise Exception(f"Config file not found: {CONFIG_FILE_PATH}")
    
    file_content = download_excel_file(file_id, token)
    wb = load_workbook(file_content, data_only=True)
    config = {}
    
    # Load therapists
    if CONFIG_SHEETS['therapists'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['therapists']]
        config['therapists'] = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                therapist = dict(zip(headers, row))
                therapist['IsActive'] = str(therapist.get('IsActive', 'TRUE')).upper() == 'TRUE'
                # Parse IsTeamLeader (can be True/False boolean or 'TRUE'/'FALSE' string)
                is_leader = therapist.get('IsTeamLeader', False)
                if isinstance(is_leader, str):
                    therapist['IsTeamLeader'] = is_leader.upper() == 'TRUE'
                else:
                    therapist['IsTeamLeader'] = bool(is_leader)
                config['therapists'].append(therapist)
        
        logging.info(f"Loaded {len(config['therapists'])} therapists")
    
    # Load teams
    if CONFIG_SHEETS['teams'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['teams']]
        config['teams'] = {}
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                team = dict(zip(headers, row))
                config['teams'][team['TeamId']] = team
        
        logging.info(f"Loaded {len(config['teams'])} teams")
    
    # Load thresholds
    config['thresholds'] = {'Physio': {}, 'OT': {}}
    
    for team_type, sheet_key in [('Physio', 'thresholds_physio'), ('OT', 'thresholds_ot')]:
        if CONFIG_SHEETS[sheet_key] in wb.sheetnames:
            ws = wb[CONFIG_SHEETS[sheet_key]]
            
            for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
                if row[0] and row[0] in ['Grad', 'CA', 'Senior', 'Team Average']:
                    config['thresholds'][team_type][row[0]] = {
                        'red_below': row[1],      # Column B: Billings_Red_Below
                        'green_min': row[2],       # Column C: Billings_Green_Min
                        'green_max': row[3],       # Column D: Billings_Green_Max
                        'blue_above': row[4]       # Column E: Billings_Blue_Above
                    }
            
            # Log what was loaded
            logging.info(f"Loaded {team_type} thresholds: {list(config['thresholds'][team_type].keys())}")
    
    # Load ceased service thresholds from row 7-8
    config['ceased_thresholds'] = {}
    if CONFIG_SHEETS['thresholds_physio'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['thresholds_physio']]
        for row in ws.iter_rows(min_row=8, max_row=9, values_only=True):
            # Look for "Ceased %" specifically to avoid header
            if row[0] and 'Ceased %' in str(row[0]):
                config['ceased_thresholds'] = {
                    'blue_below': row[1],    # Column B: Blue_Below
                    'green_min': row[2],     # Column C: Green_Min  
                    'green_max': row[3],     # Column D: Green_Max
                    'red_above': row[4]      # Column E: Red_Above
                }
                logging.info(f"Loaded ceased thresholds: {config['ceased_thresholds']}")
                break
        
        if not config['ceased_thresholds']:
            logging.warning("No ceased thresholds found (looking for 'Ceased %') - using defaults")
            config['ceased_thresholds'] = {
                'blue_below': 0.025,
                'green_min': 0.025,
                'green_max': 0.04,
                'red_above': 0.04
            }
    else:
        logging.warning("Config_Thresholds_Physio sheet not found - using default ceased thresholds")
        config['ceased_thresholds'] = {
            'blue_below': 0.025,
            'green_min': 0.025,
            'green_max': 0.04,
            'red_above': 0.04
        }
    
    # Load 1-5 rating scale thresholds (for OT and Physio non-billing columns)
    config['rating_thresholds'] = []
    if CONFIG_SHEETS['thresholds_physio'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['thresholds_physio']]
        # Look for rating scale in rows 13-17 (5 rows for ratings 1-5, header in row 12)
        for row in ws.iter_rows(min_row=13, max_row=17, values_only=True):
            if row[0] and isinstance(row[0], (int, float)) and 1 <= row[0] <= 5:
                config['rating_thresholds'].append({
                    'rating': row[0],
                    'min': row[1],
                    'max': row[2],
                    'label': row[3] if len(row) > 3 else ''
                })
        
        if config['rating_thresholds']:
            logging.info(f"Loaded {len(config['rating_thresholds'])} rating thresholds")
        else:
            logging.warning("No rating thresholds found - using defaults")
            config['rating_thresholds'] = [
                {'rating': 1, 'min': 1, 'max': 1.499, 'label': 'Unsatisfactory'},
                {'rating': 2, 'min': 1.5, 'max': 2.499, 'label': 'Needs Improvement'},
                {'rating': 3, 'min': 2.5, 'max': 3.499, 'label': 'In Progress'},
                {'rating': 4, 'min': 3.5, 'max': 4.499, 'label': 'Very Good'},
                {'rating': 5, 'min': 4.5, 'max': 5, 'label': 'Excellent'}
            ]
    
    # Load KPIs
    if CONFIG_SHEETS['kpis'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['kpis']]
        config['kpis'] = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config['kpis'].append(dict(zip(headers, row)))
    
    # Load competency history (for tracking competency changes over time)
    config['competency_history'] = []
    if CONFIG_SHEETS['competency_history'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['competency_history']]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Has a name
                record = dict(zip(headers, row))
                # Ensure EffectiveDate is a date object
                effective_date = record.get('EffectiveDate')
                if effective_date:
                    # Handle both datetime and date objects
                    if hasattr(effective_date, 'date'):
                        record['EffectiveDate'] = effective_date.date()
                    elif not isinstance(effective_date, date):
                        # Try to parse string date
                        try:
                            from datetime import datetime as dt
                            record['EffectiveDate'] = dt.strptime(str(effective_date), '%Y-%m-%d').date()
                        except:
                            logging.warning(f"Could not parse date for {record.get('Name')}: {effective_date}")
                            continue
                    config['competency_history'].append(record)
        
        logging.info(f"Loaded {len(config['competency_history'])} competency history records")
    else:
        logging.info("No Config_Competency_History sheet found - using current competencies only")
    
    # Load team average threshold history (for tracking team benchmark changes over time)
    config['team_ave_thresholds'] = []
    if CONFIG_SHEETS.get('team_ave_thresholds') and CONFIG_SHEETS['team_ave_thresholds'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['team_ave_thresholds']]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Has a team name
                record = dict(zip(headers, row))
                # Ensure EffectiveDate is a date object
                effective_date = record.get('EffectiveDate')
                if effective_date:
                    # Handle both datetime and date objects
                    if hasattr(effective_date, 'date'):
                        record['EffectiveDate'] = effective_date.date()
                    elif not isinstance(effective_date, date):
                        # Try to parse string date
                        try:
                            from datetime import datetime as dt
                            record['EffectiveDate'] = dt.strptime(str(effective_date), '%Y-%m-%d').date()
                        except:
                            logging.warning(f"Could not parse date for team {record.get('Team')}: {effective_date}")
                            continue
                    config['team_ave_thresholds'].append(record)
        
        logging.info(f"Loaded {len(config['team_ave_thresholds'])} team average threshold records")
    else:
        logging.info("No Config_TeamAve_Thresholds sheet found - using static Team Average thresholds")
    
    # Load colours
    if CONFIG_SHEETS['colours'] in wb.sheetnames:
        ws = wb[CONFIG_SHEETS['colours']]
        config['colours'] = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config['colours'][row[0]] = row[1]
    else:
        config['colours'] = DEFAULT_COLORS
    
    wb.close()
    return config


# =============================================================================
# COMPETENCY HISTORY HELPERS
# =============================================================================

# Month name to number mapping
MONTH_NAME_TO_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'June': 6,
    'July': 7, 'Aug': 8, 'Sept': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}


def extract_year_from_filename(file_path):
    """
    Extract year from filename like 'Team_Leader_2026.xlsx'.
    
    Args:
        file_path: Full file path or filename
        
    Returns:
        int: Year extracted from filename, or current year if not found
    """
    import re
    filename = file_path.split('/')[-1]  # Get just the filename
    
    # Look for 4-digit year pattern
    match = re.search(r'(\d{4})', filename)
    if match:
        year = int(match.group(1))
        # Sanity check - should be reasonable year
        if 2020 <= year <= 2100:
            logging.info(f"Extracted year {year} from filename: {filename}")
            return year
    
    # Fallback to current year
    current_year = datetime.now().year
    logging.warning(f"Could not extract year from '{filename}' - using current year {current_year}")
    return current_year


def get_competency_for_month(therapist_name, month_name, year, config):
    """
    Get the competency that was active for a therapist in a specific month.
    
    Args:
        therapist_name: Name of the therapist
        month_name: Month name ('Jan', 'Feb', etc.)
        year: Year (e.g., 2026)
        config: Config dict containing 'competency_history' and 'therapists'
        
    Returns:
        str: Competency level ('Grad', 'CA', 'Senior') or None if not found
    """
    # Get month number
    month_num = MONTH_NAME_TO_NUM.get(month_name)
    if not month_num:
        logging.warning(f"Unknown month name: {month_name}")
        return None
    
    # Create target date (mid-month to avoid edge cases)
    target_date = date(year, month_num, 15)
    
    # Get all history records for this therapist
    history = config.get('competency_history', [])
    therapist_records = [
        r for r in history 
        if r.get('Name', '').strip().lower() == therapist_name.strip().lower()
    ]
    
    if therapist_records:
        # Sort by date descending (most recent first)
        therapist_records.sort(key=lambda r: r.get('EffectiveDate', date.min), reverse=True)
        
        # Find the most recent record where EffectiveDate <= target_date
        for record in therapist_records:
            effective_date = record.get('EffectiveDate')
            if effective_date and effective_date <= target_date:
                return record.get('Competency')
    
    # No history found - fall back to current competency from Config_Therapists
    for therapist in config.get('therapists', []):
        if therapist.get('Name', '').strip().lower() == therapist_name.strip().lower():
            return therapist.get('Competency')
    
    return None


def get_therapist_current_competency(therapist_name, config):
    """
    Get the current competency from Config_Therapists (fallback/default).
    
    Args:
        therapist_name: Name of the therapist
        config: Config dict
        
    Returns:
        str: Competency level or None
    """
    for therapist in config.get('therapists', []):
        if therapist.get('Name', '').strip().lower() == therapist_name.strip().lower():
            return therapist.get('Competency')
    return None


def get_team_ave_thresholds_for_month(team_name, month_name, year, config):
    """
    Get the team average billing thresholds that were active for a specific month.
    
    This allows team benchmarks to change over time as team composition changes
    (e.g., when new grads join and lower the weighted average expectation).
    
    Args:
        team_name: 'Physio_North' | 'Physio_South' | 'OT'
        month_name: Month name ('Jan', 'Feb', etc.)
        year: Year (e.g., 2026)
        config: Config dict containing 'team_ave_thresholds'
        
    Returns:
        dict: Threshold values {'red_below', 'green_min', 'green_max', 'blue_above'}
              or None if no historical thresholds found (fall back to static)
    """
    # Get month number
    month_num = MONTH_NAME_TO_NUM.get(month_name)
    if not month_num:
        logging.warning(f"Unknown month name: {month_name}")
        return None
    
    # Create target date (mid-month to avoid edge cases)
    target_date = date(year, month_num, 15)
    
    # Get all history records for this team
    history = config.get('team_ave_thresholds', [])
    team_records = [
        r for r in history 
        if r.get('Team', '').strip() == team_name.strip()
    ]
    
    if not team_records:
        return None  # No history - fall back to static thresholds
    
    # Sort by date descending (most recent first)
    team_records.sort(key=lambda r: r.get('EffectiveDate', date.min), reverse=True)
    
    # Find the most recent record where EffectiveDate <= target_date
    for record in team_records:
        effective_date = record.get('EffectiveDate')
        if effective_date and effective_date <= target_date:
            return {
                'red_below': record.get('Billings_Red_Below'),
                'green_min': record.get('Billings_Green_Min'),
                'green_max': record.get('Billings_Green_Max'),
                'blue_above': record.get('Billings_Blue_Above')
            }
    
    return None  # No applicable record found


# =============================================================================
# KPI DASHBOARD DATA LOADER (NEW - Phase 1)
# =============================================================================

# Month columns for Jan-Dec structure
MONTH_COLUMNS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec']

# Table configurations for new Team Leader structure
PHYSIO_NORTH_TABLES = {
    'Billings_North': 'BillingsKPI',
    'Ceased_North': 'Ceased Services',
    'Documentation_North': 'Documentation',
    'Admin_North': 'Admin',
    'Attitude_North': 'Attitude'
}

PHYSIO_SOUTH_TABLES = {
    'Billings_South': 'BillingsKPI',
    'Ceased_South': 'Ceased Services',
    'Documentation_South': 'Documentation',
    'Admin_South': 'Admin',
    'Attitude_South': 'Attitude'
}

OT_TABLES = {
    'Billings_OT': 'BillingsKPI',
    'Compliance_OT': 'Compliance',
    'ReferrerEng_OT': 'Referrer Engagement',
    'Capacity_OT': 'Capacity',
    'Attitude_OT': 'Attitude'
}

SHEET_CONFIG = {
    'KPI Dashboard North': {
        'team_name': 'Physio_North',
        'tables': PHYSIO_NORTH_TABLES
    },
    'KPI Dashboard South': {
        'team_name': 'Physio_South',
        'tables': PHYSIO_SOUTH_TABLES
    },
    'KPI Dashboard OT': {
        'team_name': 'OT',
        'tables': OT_TABLES
    }
}


def read_table_data(ws, table_name, kpi_column_name):
    """Read data from one Excel table."""
    result = {}
    
    try:
        if table_name not in ws.tables:
            logging.warning(f"Table '{table_name}' not found in worksheet '{ws.title}'")
            return result
        
        table = ws.tables[table_name]
        table_range = table.ref
        
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        
        # Read header row
        header_row = []
        for col in range(min_col, max_col + 1):
            cell_value = ws.cell(row=min_row, column=col).value
            header_row.append(cell_value)
        
        # Find month column indices
        month_indices = {}
        for i, header in enumerate(header_row):
            if header in MONTH_COLUMNS:
                month_indices[header] = i
        
        # Read data rows
        for row_idx in range(min_row + 1, max_row + 1):
            name_cell = ws.cell(row=row_idx, column=min_col)
            therapist_name = name_cell.value
            
            if not therapist_name:
                continue
            
            therapist_name = str(therapist_name).strip()
            
            if therapist_name not in result:
                result[therapist_name] = {}
            
            # Read month values
            for month in MONTH_COLUMNS:
                if month in month_indices:
                    col_offset = month_indices[month]
                    cell = ws.cell(row=row_idx, column=min_col + col_offset)
                    value = cell.value
                    result[therapist_name][month] = value
        
        logging.info(f"Read {len(result)} therapists from table '{table_name}'")
        
    except Exception as e:
        logging.error(f"Error reading table '{table_name}': {str(e)}")
    
    return result


def process_dashboard_sheet(ws, team_name, table_config):
    """Extract all KPI data from one dashboard sheet."""
    logging.info(f"Processing sheet '{ws.title}' for team '{team_name}'")
    
    # Read all tables
    kpi_data = {}
    for table_name, kpi_column_name in table_config.items():
        table_data = read_table_data(ws, table_name, kpi_column_name)
        kpi_data[kpi_column_name] = table_data
    
    # Transform to therapist-centric structure
    therapist_data = {}
    
    all_therapists = set()
    for kpi_name, therapist_dict in kpi_data.items():
        all_therapists.update(therapist_dict.keys())
    
    for therapist_name in all_therapists:
        therapist_data[therapist_name] = {}
        
        for month in MONTH_COLUMNS:
            therapist_data[therapist_name][month] = {}
            
            for kpi_name, therapist_dict in kpi_data.items():
                if therapist_name in therapist_dict and month in therapist_dict[therapist_name]:
                    value = therapist_dict[therapist_name][month]
                    therapist_data[therapist_name][month][kpi_name] = value
                else:
                    therapist_data[therapist_name][month][kpi_name] = None
    
    logging.info(f"Processed {len(therapist_data)} therapists from '{ws.title}'")
    return therapist_data


def transform_to_monthly_records(therapist_data, team_name):
    """Transform therapist-centric data into monthly records."""
    records = []
    
    for therapist_name, months_data in therapist_data.items():
        for month, kpi_values in months_data.items():
            record = {
                'Name': therapist_name,
                'Month': month
            }
            record.update(kpi_values)
            records.append(record)
    
    logging.info(f"Transformed {len(records)} monthly records for team '{team_name}'")
    return records


def load_kpi_dashboard_data(wb):
    """
    Load KPI data from Team Leader Dashboard tables.
    Replaces load_master_data() for new Jan-Dec table structure.
    """
    logging.info("Loading KPI data from Dashboard tables...")
    
    master_data = {}
    
    for sheet_name, config in SHEET_CONFIG.items():
        team_name = config['team_name']
        table_config = config['tables']
        
        if sheet_name not in wb.sheetnames:
            logging.warning(f"Sheet '{sheet_name}' not found in workbook")
            master_data[team_name] = []
            continue
        
        ws = wb[sheet_name]
        therapist_data = process_dashboard_sheet(ws, team_name, table_config)
        records = transform_to_monthly_records(therapist_data, team_name)
        master_data[team_name] = records
    
    total_records = sum(len(records) for records in master_data.values())
    logging.info(f"Loaded {total_records} total records from Dashboard tables")
    for team_name, records in master_data.items():
        logging.info(f"  {team_name}: {len(records)} records")
    
    return master_data


# =============================================================================
# MASTER DATA LOADING
# =============================================================================

def load_master_data(config, token):
    """
    Load master KPI data from team sheets.
    
    Auto-detects structure:
    - NEW: KPI Dashboard tables (Jan-Dec) Ã¢â€ â€™ uses load_kpi_dashboard_data()
    - OLD: FinalPowerAutomate sheets Ã¢â€ â€™ uses legacy loader
    """
    logging.info("Loading master KPI data...")
    
    file_id = resolve_file_path(CONFIG_FILE_PATH, token)
    file_content = download_excel_file(file_id, token)
    wb = load_workbook(file_content, data_only=True)
    
    # Check if new structure exists (KPI Dashboard sheets)
    has_new_structure = any(sheet in wb.sheetnames for sheet in SHEET_CONFIG.keys())
    
    if has_new_structure:
        logging.info("Detected NEW structure (KPI Dashboard tables) - using Dashboard loader")
        master_data = load_kpi_dashboard_data(wb)
        wb.close()
        return master_data
    
    # Fall back to old structure (FinalPowerAutomate sheets)
    logging.info("Using LEGACY structure (FinalPowerAutomate sheets)")
    master_data = {}
    
    for team_id, team_config in config['teams'].items():
        sheet_name = team_config['SheetName']
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            
            team_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    team_data.append(dict(zip(headers, row)))
            
            master_data[team_id] = team_data
            logging.info(f"Loaded {len(team_data)} records for {team_id}")
    
    wb.close()
    return master_data


def create_from_template(therapist, config, token):
    """Create new therapist file from template."""
    name = therapist['Name']
    team_id = therapist['Team']
    team_type = 'OT' if team_id == 'OT' else 'Physio'
    file_path = therapist['FilePath']
    
    logging.info(f"Creating new file for {name} from {team_type} template")
    
    # Determine template path (in tests folder for now)
    template_path = f"/Excel files/KPI/templates/Template_{team_type}.xlsx"
    
    # Download template
    template_file_id = resolve_file_path(template_path, token)
    if not template_file_id:
        logging.error(f"Template not found: {template_path}")
        return False
    
    try:
        # Download template file
        template_content = download_excel_file(template_file_id, token)
        wb = load_workbook(template_content)
        
        # Find Dashboard sheet
        dashboard_sheet = None
        for sheet_name in wb.sheetnames:
            if 'Dashboard' in sheet_name and 'KPI' not in sheet_name:
                dashboard_sheet = sheet_name
                break
        
        if not dashboard_sheet:
            logging.error(f"Dashboard sheet not found in template")
            wb.close()
            return False
        
        ws = wb[dashboard_sheet]
        
        # Write therapist name to B2 (merged with C2)
        ws.cell(row=2, column=2, value=name)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        wb.close()
        
        # Upload as new file
        upload_excel_file(file_path, output, token)
        logging.info(f"Created new file for {name} at {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to create file from template: {e}")
        return False


# NOTE: Skipping sync_fte_sheet, sync_team_tables functions
# These aren't needed for basic individual sheet update testing
# Include them from original file if needed


def apply_billing_formatting(ws, thresholds, colours):
    """Apply conditional formatting to billing KPI cells - Red/Green/Blue scale using config values."""
    if not thresholds:
        logging.warning("No thresholds provided for billing formatting")
        return
    
    # Get threshold values from config
    red_below = thresholds.get('red_below', 0)
    green_min = thresholds.get('green_min', 0)
    green_max = thresholds.get('green_max', 100)
    blue_above = thresholds.get('blue_above', 100)
    
    logging.info(f"Billing thresholds: red<{red_below}, green {green_min}-{green_max}, blue>{blue_above}")
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = 'FF4FC3F7'  # Blue for excellent
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    # Apply to BillingsKPI column (column C, rows 5-17 including average)
    billing_range = 'C5:C17'
    
    # Clear existing conditional formatting
    ws.conditional_formatting._cf_rules = {}
    
    # Priority order (highest to lowest):
    # 1. White for blank cells
    ws.conditional_formatting.add(billing_range,
        FormulaRule(formula=[f'=LEN(TRIM(C5))=0'], fill=white_fill))
    
    # 2. Blue for > blue_above (excellent)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='greaterThan', formula=[str(blue_above)], fill=blue_fill))
    
    # 3. Green for >= green_min (meets threshold)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(green_min)], fill=green_fill))
    
    # 4. Red for < green_min (below threshold)
    ws.conditional_formatting.add(billing_range,
        CellIsRule(operator='lessThan', formula=[str(green_min)], fill=red_fill))


def apply_ceased_services_formatting(ws, ceased_thresholds, colours):
    """Apply conditional formatting to Ceased Services column (Physio only) - Red/Green/Blue scale using config values."""
    
    # Get threshold values from config
    blue_below = ceased_thresholds.get('blue_below', 0.015)
    green_min = ceased_thresholds.get('green_min', 0.015)  
    green_max = ceased_thresholds.get('green_max', 0.04)
    red_above = ceased_thresholds.get('red_above', 0.04)
    
    logging.info(f"Ceased thresholds: blue<{blue_below}, green {green_min}-{green_max}, red>{red_above}")
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = 'FF4FC3F7'  # Blue for excellent
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    # Ceased Services: Column D, rows 5-17 (including average)
    # Lower is better
    ceased_range = 'D5:D17'
    
    # 1. White for blank
    ws.conditional_formatting.add(ceased_range,
        FormulaRule(formula=[f'=LEN(TRIM(D5))=0'], fill=white_fill))
    
    # 2. Blue for < blue_below (excellent - very low ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='lessThan', formula=[str(blue_below)], fill=blue_fill))
    
    # 3. Green for < red_above (good - acceptable ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='lessThan', formula=[str(red_above)], fill=green_fill))
    
    # 4. Red for >= red_above (poor - high ceased rate)
    ws.conditional_formatting.add(ceased_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(red_above)], fill=red_fill))


def apply_rating_scale_formatting(ws, cell_ranges, rating_thresholds, colours):
    """Apply conditional formatting using 1-5 rating scale from config. Works for both whole numbers and averages."""
    
    red_color = colours.get('red', DEFAULT_COLORS['red'])
    amber_color = colours.get('amber', DEFAULT_COLORS['amber'])
    yellow_color = 'FFFFF176'  # Yellow
    green_color = colours.get('green', DEFAULT_COLORS['green'])
    blue_color = 'FF4FC3F7'  # Blue
    white_color = colours.get('white', DEFAULT_COLORS['white'])
    
    red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
    amber_fill = PatternFill(start_color=amber_color, end_color=amber_color, fill_type='solid')
    yellow_fill = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')
    green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    white_fill = PatternFill(start_color=white_color, end_color=white_color, fill_type='solid')
    
    # Map ratings to colors
    # Rating 1: Red
    # Rating 2: Amber/Orange  
    # Rating 3: Yellow
    # Rating 4: Green
    # Rating 5: Blue
    color_map = {
        1: red_fill,
        2: amber_fill,
        3: yellow_fill,
        4: green_fill,
        5: blue_fill
    }
    
    for cell_range in cell_ranges:
        col_letter = cell_range.split(':')[0][0]  # Extract column letter
        
        # 1. White for blank
        ws.conditional_formatting.add(cell_range,
            FormulaRule(formula=[f'=LEN(TRIM({col_letter}5))=0'], fill=white_fill))
        
        # Apply rules from config (highest rating first for priority)
        for threshold in reversed(rating_thresholds):  # Start with 5, work down to 1
            rating = threshold['rating']
            min_val = threshold['min']
            max_val = threshold['max']
            fill = color_map.get(rating, white_fill)
            
            # Use 'between' for all except the extremes
            if rating == 5:
                # >= 4.5 (for averages) or 5 (for whole numbers)
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='greaterThanOrEqual', formula=[str(min_val)], fill=fill))
            elif rating == 1:
                # < 1.5 (for averages) or 1 (for whole numbers)
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='lessThan', formula=[str(max_val)], fill=fill))
            else:
                # Between min and max
                ws.conditional_formatting.add(cell_range,
                    CellIsRule(operator='between', formula=[str(min_val), str(max_val)], fill=fill))


# =============================================================================
# INDIVIDUAL SHEET UPDATE
# =============================================================================

def update_individual_sheet(therapist, config, master_data, token):
    """Update a single therapist's individual KPI sheet."""
    name = therapist['Name']
    team_id = therapist['Team']
    team_type = 'OT' if team_id == 'OT' else 'Physio'
    competency = therapist.get('Competency', 'CA')
    file_path = therapist['FilePath']
    
    logging.info(f"Updating {name}...")
    
    file_id = resolve_file_path(file_path, token)
    
    if not file_id:
        # File doesn't exist - try creating from template
        logging.warning(f"File not found: {file_path} - attempting to create from template")
        if create_from_template(therapist, config, token):
            # Template created successfully - try resolving file_id again
            file_id = resolve_file_path(file_path, token)
            if not file_id:
                logging.error(f"File still not found after template creation: {file_path}")
                return False
        else:
            logging.error(f"Failed to create file from template: {file_path}")
            return False
    
    team_data = master_data.get(team_id, [])
    
    # Match therapist records - try both 'Name' (new) and 'UniqueTeamMembers' (old)
    kpi_records = [r for r in team_data if r.get('Name') == name or r.get('UniqueTeamMembers') == name]
    
    if not kpi_records:
        logging.warning(f"No KPI data for {name} in team {team_id} - but continuing to update thresholds")
        kpi_records = []
    
    file_content = download_excel_file(file_id, token)
    wb = load_workbook(file_content)
    
    dashboard_sheet = None
    for sheet_name in wb.sheetnames:
        if 'Dashboard' in sheet_name and 'KPI' not in sheet_name:
            dashboard_sheet = sheet_name
            break
    
    if not dashboard_sheet:
        wb.close()
        return False
    
    ws = wb[dashboard_sheet]
    
    # Jan-Dec month mapping (rows 5-16)
    month_rows = {
        'Jan': 5, 'January': 5,
        'Feb': 6, 'February': 6,
        'Mar': 7, 'March': 7,
        'Apr': 8, 'April': 8,
        'May': 9,
        'Jun': 10, 'June': 10,
        'Jul': 11, 'July': 11,
        'Aug': 12, 'August': 12,
        'Sep': 13, 'Sept': 13, 'September': 13,
        'Oct': 14, 'October': 14,
        'Nov': 15, 'November': 15,
        'Dec': 16, 'December': 16
    }
    
    if team_type == 'OT':
        kpi_columns = {
            'BillingsKPI': 3,  # Column C
            'Compliance': 4,  # Column D (NEW: matches Dashboard table name)
            'Referrer Engagement': 5,  # Column E (NEW: matches Dashboard table name)
            'Capacity': 6,  # Column F (NEW: matches Dashboard table name)
            'Attitude': 7  # Column G (NEW: matches Dashboard table name)
        }
    else:
        kpi_columns = {
            'BillingsKPI': 3,  # Column C
            'Ceased Services': 4,  # Column D
            'Documentation': 5,  # Column E
            'Admin': 6,  # Column F (NEW: matches Dashboard table name)
            'Attitude': 7  # Column G (NEW: matches Dashboard table name)
        }
    
    for record in kpi_records:
        month = record.get('Month')
        if month not in month_rows:
            continue
        row = month_rows[month]
        for kpi_name, col in kpi_columns.items():
            value = record.get(kpi_name)
            if value is not None:
                ws.cell(row=row, column=col, value=value)
    
    thresholds_all = config['thresholds'].get(team_type, {})
    
    # Populate ALL three competency rows (Grad, CA, Senior) - rows 21-23
    competency_rows = {'Grad': 21, 'CA': 22, 'Senior': 23}
    
    for comp_level, row in competency_rows.items():
        comp_thresholds = thresholds_all.get(comp_level, {})
        if comp_thresholds:
            try:
                # Display format: Below | Good | Excellent
                # Column C: < green_min (red/below)
                # Column D: green_min - green_max (green/good range)
                # Column E: > blue_above (blue/excellent)
                ws.cell(row=row, column=3, value=f"<{comp_thresholds.get('green_min', '')}")
                ws.cell(row=row, column=4, value=f"{comp_thresholds.get('green_min', '')}-{comp_thresholds.get('green_max', '')}")
                ws.cell(row=row, column=5, value=f">{comp_thresholds.get('blue_above', '')}")
            except Exception as e:
                logging.warning(f"Could not write {comp_level} thresholds to row {row} (possibly merged cells): {e}")
    
    # Get thresholds for THIS therapist's competency level (for conditional formatting)
    thresholds = thresholds_all.get(competency, {})
    
    # Populate B26:B28 - Ceased Services thresholds (from config)
    ceased = config.get('ceased_thresholds', {})
    if team_type == 'Physio':  # Only for Physio
        # Use values from config (already loaded with proper defaults)
        blue_below = ceased.get('blue_below', 0.015)
        green_min = ceased.get('green_min', 0.015)
        green_max = ceased.get('green_max', 0.04)
        red_above = ceased.get('red_above', 0.04)
        
        try:
            # Convert to percentages for display
            # Row 26: Blue (Excellent) - < blue_below
            # Row 27: Green (Good) - green_min to green_max  
            # Row 28: Red (Needs Improvement) - >= red_above
            ws.cell(row=26, column=2, value=f"<{blue_below*100:.1f}%")
            ws.cell(row=27, column=2, value=f"{green_min*100:.1f}-{green_max*100:.1f}%")
            ws.cell(row=28, column=2, value=f">{red_above*100:.1f}%")
        except Exception as e:
            logging.warning(f"Could not write ceased thresholds to rows 26-28 (possibly merged cells): {e}")
    
    # Apply conditional formatting
    if team_type == 'OT':
        # OT: BillingsKPI (C) uses thresholds, columns D-G use 1-5 rating scale
        apply_billing_formatting(ws, thresholds, config['colours'])
        
        # Apply rating scale to columns D-G (rows 5-17, including average)
        ot_ranges = ['D5:D17', 'E5:E17', 'F5:F17', 'G5:G17']
        apply_rating_scale_formatting(ws, ot_ranges, config['rating_thresholds'], config['colours'])
        
    else:
        # Physio: BillingsKPI (C) and Ceased Services (D) use thresholds, columns E-G use 1-5 rating scale
        apply_billing_formatting(ws, thresholds, config['colours'])
        apply_ceased_services_formatting(ws, config['ceased_thresholds'], config['colours'])
        
        # Apply rating scale to columns E-G (rows 5-17, including average)
        physio_ranges = ['E5:E17', 'F5:F17', 'G5:G17']
        apply_rating_scale_formatting(ws, physio_ranges, config['rating_thresholds'], config['colours'])
    
    output = BytesIO()
    wb.save(output)
    wb.close()
    
    upload_excel_file(file_path, output, token)
    logging.info(f"Updated {name}")
    return True


# NOTE: Skipping archive_year, kpi_sync_function, HTTP triggers
# These are Azure Function specific and not needed for local testing

print("✅ function_app_local.py loaded - ready for local testing")
