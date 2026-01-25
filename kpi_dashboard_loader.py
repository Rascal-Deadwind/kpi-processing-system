"""
KPI Dashboard Data Loader - Phase 1
Reads KPI data from Team Leader Dashboard tables (Jan-Dec structure)
Replaces load_master_data() which read from FinalPowerAutomate sheets
"""

import logging
from openpyxl import load_workbook

# =============================================================================
# TABLE CONFIGURATION
# =============================================================================

# Map table names to output column names
PHYSIO_NORTH_TABLES = {
    'Billings_North': 'BillingsKPI',
    'Ceased_North': 'Ceased Services',
    'Documentation_North': 'Documentation',
    'Admin_North': 'Admin',
    'Attitude_North': 'Attitude',
    'Average_North': 'Team Average'  # ADD: Average table
}

PHYSIO_SOUTH_TABLES = {
    'Billings_South': 'BillingsKPI',
    'Ceased_South': 'Ceased Services',
    'Documentation_South': 'Documentation',
    'Admin_South': 'Admin',
    'Attitude_South': 'Attitude',
    'Average_South': 'Team Average'  # ADD: Average table
}

OT_TABLES = {
    'Billings_OT': 'BillingsKPI',
    'Compliance_OT': 'Compliance',
    'ReferrerEng_OT': 'Referrer Engagement',
    'Capacity_OT': 'Capacity',
    'Attitude_OT': 'Attitude',
    'Average_OT': 'Team Average'  # ADD: Average table
}

# Sheet configuration
# MMP Dashboard tables (organization-wide averages)
MMP_TABLES_NORTH = {
    'Average_North14': 'Team Average (North)'
}

MMP_TABLES_SOUTH = {
    'Average_South10': 'Team Average (South)'
}

MMP_TABLES_OT = {
    'Average_OT20': 'Team Average (OT)'
}

# Sheet configuration
SHEET_CONFIG = {
    'KPI Dashboard North': {
        'team_name': 'Physio_North',
        'tables': PHYSIO_NORTH_TABLES,
        'average_tables': ['Average_North']  # Mark which tables are averages
    },
    'KPI Dashboard South': {
        'team_name': 'Physio_South',
        'tables': PHYSIO_SOUTH_TABLES,
        'average_tables': ['Average_South']
    },
    'KPI Dashboard OT': {
        'team_name': 'OT',
        'tables': OT_TABLES,
        'average_tables': ['Average_OT']
    },
    'MMP Dashboard': {
        'team_name': 'MMP',
        'is_average_sheet': True,  # Entire sheet is averages
        'tables': {
            **MMP_TABLES_NORTH,
            **MMP_TABLES_SOUTH,
            **MMP_TABLES_OT
        }
    }
}

# Month columns (C-N in tables = Jan-Dec, plus Average column)
MONTH_COLUMNS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec', 'Average']


# =============================================================================
# STEP 1: READ TABLE DATA
# =============================================================================

def read_table_data(ws, table_name, kpi_column_name):
    """
    Read data from one Excel table.
    
    Args:
        ws: openpyxl worksheet object
        table_name: Name of the table (e.g., 'Billings_North')
        kpi_column_name: Target column name for output (e.g., 'BillingsKPI')
        
    Returns:
        dict: {therapist_name: {month: value, ...}, ...}
        
    Example:
        {
            'Chris': {'Jan': 5.2, 'Feb': 5.4, 'Mar': None, ...},
            'Sarah': {'Jan': 6.1, 'Feb': 6.3, ...},
            ...
        }
    """
    result = {}
    
    try:
        # Get the table object
        if table_name not in ws.tables:
            logging.warning(f"Table '{table_name}' not found in worksheet '{ws.title}'")
            return result
        
        table = ws.tables[table_name]
        table_range = table.ref
        
        # Parse the table range (e.g., "B13:O23")
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        
        # Read header row to get month positions
        # Assuming: Column 0 = Name, Columns 1-12 = Jan-Dec, Column 13 = Average
        header_row = []
        for col in range(min_col, max_col + 1):
            cell_value = ws.cell(row=min_row, column=col).value
            header_row.append(cell_value)
        
        logging.debug(f"Table '{table_name}' header: {header_row[:5]}...")
        
        # Find month column indices (should be columns 1-12 after Name column)
        month_indices = {}
        for i, header in enumerate(header_row):
            if header in MONTH_COLUMNS:
                month_indices[header] = i
        
        # Read data rows (skip header)
        for row_idx in range(min_row + 1, max_row + 1):
            # Get therapist name (first column)
            name_cell = ws.cell(row=row_idx, column=min_col)
            therapist_name = name_cell.value
            
            if not therapist_name:
                continue  # Skip empty rows
            
            # Clean up name (strip whitespace)
            therapist_name = str(therapist_name).strip()
            
            # Initialize therapist data
            if therapist_name not in result:
                result[therapist_name] = {}
            
            # Read month values
            for month in MONTH_COLUMNS:
                if month in month_indices:
                    col_offset = month_indices[month]
                    cell = ws.cell(row=row_idx, column=min_col + col_offset)
                    value = cell.value
                    
                    # Store value (None if empty, otherwise the actual value)
                    result[therapist_name][month] = value
        
        logging.info(f"Read {len(result)} therapists from table '{table_name}'")
        
    except Exception as e:
        logging.error(f"Error reading table '{table_name}': {str(e)}")
        import traceback
        traceback.print_exc()
    
    return result


# =============================================================================
# STEP 2: PROCESS DASHBOARD SHEET
# =============================================================================

def process_dashboard_sheet(ws, team_name, table_config):
    """
    Extract all KPI data from one dashboard sheet.
    
    Args:
        ws: openpyxl worksheet object
        team_name: 'Physio_North' | 'Physio_South' | 'OT'
        table_config: Dict mapping table names to KPI column names
        
    Returns:
        dict: {
            therapist_name: {
                month: {kpi_name: value, ...},
                ...
            },
            ...
        }
        
    Example:
        {
            'Chris': {
                'Jan': {'BillingsKPI': 5.2, 'Ceased Services': 0.03, ...},
                'Feb': {'BillingsKPI': 5.4, ...},
                ...
            }
        }
    """
    logging.info(f"Processing sheet '{ws.title}' for team '{team_name}'")
    
    # Read all tables for this sheet
    kpi_data = {}  # {kpi_name: {therapist: {month: value}}}
    
    for table_name, kpi_column_name in table_config.items():
        table_data = read_table_data(ws, table_name, kpi_column_name)
        kpi_data[kpi_column_name] = table_data
    
    # Transform from KPI-centric to therapist-centric structure
    therapist_data = {}
    
    # Get all unique therapist names across all KPIs
    all_therapists = set()
    for kpi_name, therapist_dict in kpi_data.items():
        all_therapists.update(therapist_dict.keys())
    
    # Build therapist-centric structure
    for therapist_name in all_therapists:
        therapist_data[therapist_name] = {}
        
        # For each month
        for month in MONTH_COLUMNS:
            therapist_data[therapist_name][month] = {}
            
            # For each KPI, get the value for this therapist/month
            for kpi_name, therapist_dict in kpi_data.items():
                if therapist_name in therapist_dict and month in therapist_dict[therapist_name]:
                    value = therapist_dict[therapist_name][month]
                    therapist_data[therapist_name][month][kpi_name] = value
                else:
                    therapist_data[therapist_name][month][kpi_name] = None
    
    logging.info(f"Processed {len(therapist_data)} therapists from '{ws.title}'")
    
    return therapist_data


# =============================================================================
# STEP 3: TRANSFORM TO MONTHLY RECORDS
# =============================================================================

def transform_to_monthly_records(therapist_data, team_name):
    """
    Transform therapist-centric data into monthly records.
    
    Args:
        therapist_data: {
            therapist_name: {
                month: {kpi_name: value, ...},
                ...
            }
        }
        team_name: 'Physio_North' | 'Physio_South' | 'OT'
        
    Returns:
        list: [
            {'Name': 'Chris', 'Month': 'Jan', 'BillingsKPI': 5.2, ...},
            {'Name': 'Chris', 'Month': 'Feb', ...},
            ...
        ]
    """
    records = []
    
    for therapist_name, months_data in therapist_data.items():
        for month, kpi_values in months_data.items():
            # Build one record for this therapist/month
            record = {
                'Name': therapist_name,
                'Month': month
            }
            
            # Add all KPI values
            record.update(kpi_values)
            
            records.append(record)
    
    logging.info(f"Transformed {len(records)} monthly records for team '{team_name}'")
    
    return records


# =============================================================================
# STEP 4: MAIN LOADER FUNCTION
# =============================================================================

def load_kpi_dashboard_data(wb):
    """
    Load KPI data from Team Leader Dashboard tables.
    
    This replaces load_master_data() which read from FinalPowerAutomate sheets.
    
    Args:
        wb: openpyxl workbook object (already loaded)
        
    Returns:
        dict: {
            'Physio_North': [
                {'Name': 'Chris', 'Month': 'Jan', 'BillingsKPI': 5.2, ...},
                ...
            ],
            'Physio_South': [...],
            'OT': [...]
        }
    """
    logging.info("Loading KPI data from Dashboard tables...")
    
    master_data = {}
    
    # Process each sheet
    for sheet_name, config in SHEET_CONFIG.items():
        team_name = config['team_name']
        table_config = config['tables']
        
        if sheet_name not in wb.sheetnames:
            logging.warning(f"Sheet '{sheet_name}' not found in workbook")
            master_data[team_name] = []
            continue
        
        # Get worksheet
        ws = wb[sheet_name]
        
        # Process sheet to get therapist-centric data
        therapist_data = process_dashboard_sheet(ws, team_name, table_config)
        
        # Transform to monthly records format
        records = transform_to_monthly_records(therapist_data, team_name)
        
        master_data[team_name] = records
    
    # Log summary
    total_records = sum(len(records) for records in master_data.values())
    logging.info(f"Loaded {total_records} total records from Dashboard tables")
    for team_name, records in master_data.items():
        logging.info(f"  {team_name}: {len(records)} records")
    
    return master_data


# =============================================================================
# STANDALONE TEST FUNCTION
# =============================================================================

def test_loader(file_path):
    """Test the loader with a local file."""
    print("="*70)
    print("Testing KPI Dashboard Loader")
    print("="*70)
    
    wb = load_workbook(file_path, data_only=True)
    
    print(f"\nWorkbook sheets: {wb.sheetnames}")
    
    # Test reading one table
    print("\n" + "="*70)
    print("Test 1: Read single table (Billings_North)")
    print("="*70)
    
    if 'KPI Dashboard North' in wb.sheetnames:
        ws = wb['KPI Dashboard North']
        table_data = read_table_data(ws, 'Billings_North', 'BillingsKPI')
        print(f"âœ… Found {len(table_data)} therapists")
        for name, months in list(table_data.items())[:2]:
            print(f"  {name}: Jan={months.get('Jan')}, Feb={months.get('Feb')}")
    
    # Test full load
    print("\n" + "="*70)
    print("Test 2: Full data load")
    print("="*70)
    
    master_data = load_kpi_dashboard_data(wb)
    
    print(f"\nâœ… Loaded data for {len(master_data)} teams")
    for team_name, records in master_data.items():
        print(f"  {team_name}: {len(records)} records")
        if records:
            print(f"    Sample: {records[0]}")
    
    wb.close()
    
    print("\n" + "="*70)
    print("Test Complete!")
    print("="*70)


if __name__ == "__main__":
    import sys
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    if len(sys.argv) > 1:
        test_loader(sys.argv[1])
    else:
        print("Usage: python kpi_dashboard_loader.py <path_to_excel_file>")
