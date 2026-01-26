"""
Phase 2: Team Leader Table Formatting
Apply competency-based conditional formatting to Team Leader KPI Dashboard tables

FIXES (Nov 30, 2025):
1. Average table billing rows now use 'CA' competency (not 'Team Average')
2. Ceased thresholds default to 0.025 (not 0.015)
3. Rating scale properly applies blue for >=4.5
4. Rating/Ceased tables only format therapists in Config_Therapists
5. Ceased Services tables use correct 0.025 threshold
"""

import logging
from openpyxl.styles import PatternFill
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList


# =============================================================================
# CONFIGURATION
# =============================================================================

DATA_COLUMNS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec', 'Average']

# Month name to number mapping for competency history
MONTH_TO_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'June': 6,
    'July': 7, 'Aug': 8, 'Sept': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

DEFAULT_COLORS = {
    'red': 'FFE47373',
    'amber': 'FFFFB74D',
    'yellow': 'FFFFF176',
    'green': 'FF81C784',
    'blue': 'FF4FC3F7',
    'white': 'FFFFFFFF'
}

DEFAULT_CEASED_THRESHOLDS = {
    'blue_below': 0.025,
    'green_min': 0.025,
    'green_max': 0.04,
    'red_above': 0.04
}

DEFAULT_RATING_THRESHOLDS = [
    {'rating': 1, 'min': 1.0, 'max': 1.5},
    {'rating': 2, 'min': 1.5, 'max': 2.5},
    {'rating': 3, 'min': 2.5, 'max': 3.5},
    {'rating': 4, 'min': 3.5, 'max': 4.5},
    {'rating': 5, 'min': 4.5, 'max': 5.0}
]

SHEET_CONFIG = {
    'KPI Dashboard North': {
        'team_name': 'Physio_North',
        'team_type': 'Physio',
        'tables': {
            'Billings_North': {'kpi_type': 'billing'},
            'Ceased_North': {'kpi_type': 'ceased'},
            'Documentation_North': {'kpi_type': 'rating'},
            'Admin_North': {'kpi_type': 'rating'},
            'Attitude_North': {'kpi_type': 'rating'},
            'Average_North': {'kpi_type': 'average', 'is_average': True},
        }
    },
    'KPI Dashboard South': {
        'team_name': 'Physio_South',
        'team_type': 'Physio',
        'tables': {
            'Billings_South': {'kpi_type': 'billing'},
            'Ceased_South': {'kpi_type': 'ceased'},
            'Documentation_South': {'kpi_type': 'rating'},
            'Admin_South': {'kpi_type': 'rating'},
            'Attitude_South': {'kpi_type': 'rating'},
            'Average_South': {'kpi_type': 'average', 'is_average': True},
        }
    },
    'KPI Dashboard OT': {
        'team_name': 'OT',
        'team_type': 'OT',
        'tables': {
            'Billings_OT': {'kpi_type': 'billing'},
            'Compliance_OT': {'kpi_type': 'rating'},
            'ReferrerEng_OT': {'kpi_type': 'rating'},
            'Capacity_OT': {'kpi_type': 'rating'},
            'Attitude_OT': {'kpi_type': 'rating'},
            'Average_OT': {'kpi_type': 'average', 'is_average': True},
        }
    },
    'MMP Dashboard': {
        'team_name': 'MMP',
        'team_type': None,
        'is_mmp': True,
        'tables': {
            'Average_North14': {'kpi_type': 'mmp_average', 'team_type': 'Physio'},
            'Average_South10': {'kpi_type': 'mmp_average', 'team_type': 'Physio'},
            'Average_OT20': {'kpi_type': 'mmp_average', 'team_type': 'OT'},
        }
    }
}


# =============================================================================
# HELPERS
# =============================================================================

def build_competency_map(config, team_name):
    """Map therapist names to competencies for a specific team."""
    competency_map = {}
    for therapist in config.get('therapists', []):
        if therapist.get('Team') == team_name:
            name = therapist.get('Name')
            competency = therapist.get('Competency')
            if name and competency:
                competency_map[name.strip()] = competency
    logging.info(f"Competency map for {team_name}: {len(competency_map)} therapists - {list(competency_map.keys())}")
    return competency_map


def create_fills(colours):
    """Create PatternFill objects from config colours."""
    def get_color(name):
        return colours.get(name, DEFAULT_COLORS.get(name, 'FFFFFFFF'))
    
    return {
        'red': PatternFill(start_color=get_color('red'), end_color=get_color('red'), fill_type='solid'),
        'amber': PatternFill(start_color=get_color('amber'), end_color=get_color('amber'), fill_type='solid'),
        'yellow': PatternFill(start_color=get_color('yellow'), end_color=get_color('yellow'), fill_type='solid'),
        'green': PatternFill(start_color=get_color('green'), end_color=get_color('green'), fill_type='solid'),
        'blue': PatternFill(start_color=get_color('blue'), end_color=get_color('blue'), fill_type='solid'),
        'white': PatternFill(start_color=get_color('white'), end_color=get_color('white'), fill_type='solid'),
    }


def get_table_info(ws, table_name):
    """Get table boundaries and column mapping."""
    if table_name not in ws.tables:
        return None
    
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    header_row = [ws.cell(row=min_row, column=col).value for col in range(min_col, max_col + 1)]
    
    data_cols = {}
    for i, header in enumerate(header_row):
        if header in DATA_COLUMNS:
            data_cols[header] = min_col + i
    
    return {
        'min_col': min_col,
        'min_row': min_row,
        'max_col': max_col,
        'max_row': max_row,
        'data_cols': data_cols
    }


def detect_row_kpi_type(row_label):
    """Detect KPI type from row label (for average tables)."""
    if not row_label:
        return 'rating'
    label_upper = str(row_label).upper()
    if 'BILLING' in label_upper:
        return 'billing'
    if 'CEASED' in label_upper:
        return 'ceased'
    return 'rating'


def clear_all_conditional_formatting(ws):
    """Completely clear all conditional formatting from a worksheet."""
    try:
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            if hasattr(ws.conditional_formatting, '_cf_rules'):
                old_count = len(ws.conditional_formatting._cf_rules)
                if old_count > 0:
                    logging.info(f"Clearing {old_count} existing rules from '{ws.title}'")
        ws.conditional_formatting = ConditionalFormattingList()
        return True
    except Exception as e:
        logging.error(f"Error clearing conditional formatting: {e}")
        return False


def extract_year_from_filename(file_path):
    """Extract year from filename like 'Team_Leader_2026.xlsx'."""
    import re
    if not file_path:
        return None
    filename = file_path.split('/')[-1]
    match = re.search(r'(\d{4})', filename)
    if match:
        year = int(match.group(1))
        if 2020 <= year <= 2100:
            return year
    return None


def get_competency_ranges_for_therapist(therapist_name, year, data_cols, config):
    """
    Get column ranges for each competency period for a therapist.
    
    Returns list of tuples: [(competency, start_col, end_col), ...]
    
    Example for Chris (CA Jan, Senior from Feb):
    [('CA', 3, 3), ('Senior', 4, 15)]  # C for Jan, D-O for Feb-Dec+Avg
    """
    from datetime import date
    
    history = config.get('competency_history', [])
    
    # Get records for this therapist
    therapist_records = [
        r for r in history
        if r.get('Name', '').strip().lower() == therapist_name.strip().lower()
        and r.get('EffectiveDate')
    ]
    
    if not therapist_records or not year:
        # No history - return None to use standard single-range formatting
        return None
    
    # Sort by date ascending
    therapist_records.sort(key=lambda r: r.get('EffectiveDate'))
    
    # Build month -> column mapping (excluding Average)
    month_cols = {m: c for m, c in data_cols.items() if m != 'Average'}
    avg_col = data_cols.get('Average')
    
    # Sort months by their calendar order
    sorted_months = sorted(month_cols.items(), key=lambda x: MONTH_TO_NUM.get(x[0], 99))
    
    # Determine competency for each month
    month_competencies = {}
    for month_name, col in sorted_months:
        month_num = MONTH_TO_NUM.get(month_name)
        if not month_num:
            continue
        
        # Target date is mid-month of that month in the given year
        target_date = date(year, month_num, 15)
        
        # Find applicable competency (latest record where EffectiveDate <= target)
        applicable_comp = None
        for record in therapist_records:
            eff_date = record.get('EffectiveDate')
            if hasattr(eff_date, 'date'):
                eff_date = eff_date.date()
            if eff_date <= target_date:
                applicable_comp = record.get('Competency')
        
        if applicable_comp:
            month_competencies[month_name] = (applicable_comp, col)
    
    if not month_competencies:
        return None
    
    # Group consecutive months with same competency into ranges
    ranges = []
    current_comp = None
    range_start_col = None
    range_end_col = None
    
    for month_name, col in sorted_months:
        if month_name not in month_competencies:
            continue
        comp, col = month_competencies[month_name]
        
        if comp == current_comp:
            # Extend current range
            range_end_col = col
        else:
            # Save previous range if exists
            if current_comp is not None:
                ranges.append((current_comp, range_start_col, range_end_col))
            # Start new range
            current_comp = comp
            range_start_col = col
            range_end_col = col
    
    # Don't forget the last range
    if current_comp is not None:
        ranges.append((current_comp, range_start_col, range_end_col))
    
    # Add Average column to the last competency's range
    if ranges and avg_col:
        last_comp, last_start, last_end = ranges[-1]
        ranges[-1] = (last_comp, last_start, avg_col)
    
    return ranges if len(ranges) > 1 else None  # Only return if there are multiple ranges


def get_team_ave_threshold_ranges(team_name, year, data_cols, config):
    """
    Get column ranges for each threshold period for a team's billing row.
    
    Similar to get_competency_ranges_for_therapist but for team averages.
    
    Returns list of tuples: [(thresholds_dict, start_col, end_col), ...]
    
    Example for OT team (thresholds change in March when grads join):
    [({'green_min': 4.3, ...}, 3, 4), ({'green_min': 4.0, ...}, 5, 15)]
    # C-D for Jan-Feb, E-O for Mar-Dec+Avg
    
    Returns None if no history exists (use static thresholds instead).
    """
    from datetime import date
    
    history = config.get('team_ave_thresholds', [])
    
    # Get records for this team
    team_records = [
        r for r in history
        if r.get('Team', '').strip() == team_name.strip()
        and r.get('EffectiveDate')
    ]
    
    if not team_records or not year:
        # No history - return None to use standard single-range formatting
        return None
    
    # Sort by date ascending
    team_records.sort(key=lambda r: r.get('EffectiveDate'))
    
    # Build month -> column mapping (excluding Average)
    month_cols = {m: c for m, c in data_cols.items() if m != 'Average'}
    avg_col = data_cols.get('Average')
    
    # Sort months by their calendar order
    sorted_months = sorted(month_cols.items(), key=lambda x: MONTH_TO_NUM.get(x[0], 99))
    
    # Determine thresholds for each month
    month_thresholds = {}
    for month_name, col in sorted_months:
        month_num = MONTH_TO_NUM.get(month_name)
        if not month_num:
            continue
        
        # Target date is mid-month of that month in the given year
        target_date = date(year, month_num, 15)
        
        # Find applicable thresholds (latest record where EffectiveDate <= target)
        applicable_thresholds = None
        for record in team_records:
            eff_date = record.get('EffectiveDate')
            if hasattr(eff_date, 'date'):
                eff_date = eff_date.date()
            if eff_date <= target_date:
                applicable_thresholds = {
                    'red_below': record.get('Billings_Red_Below'),
                    'green_min': record.get('Billings_Green_Min'),
                    'green_max': record.get('Billings_Green_Max'),
                    'blue_above': record.get('Billings_Blue_Above')
                }
        
        if applicable_thresholds:
            # Create a hashable key for grouping (based on threshold values)
            threshold_key = (
                applicable_thresholds.get('green_min'),
                applicable_thresholds.get('blue_above')
            )
            month_thresholds[month_name] = (applicable_thresholds, threshold_key, col)
    
    if not month_thresholds:
        return None
    
    # Group consecutive months with same thresholds into ranges
    ranges = []
    current_key = None
    current_thresholds = None
    range_start_col = None
    range_end_col = None
    
    for month_name, col in sorted_months:
        if month_name not in month_thresholds:
            continue
        thresholds, threshold_key, col = month_thresholds[month_name]
        
        if threshold_key == current_key:
            # Extend current range
            range_end_col = col
        else:
            # Save previous range if exists
            if current_thresholds is not None:
                ranges.append((current_thresholds, range_start_col, range_end_col))
            # Start new range
            current_key = threshold_key
            current_thresholds = thresholds
            range_start_col = col
            range_end_col = col
    
    # Don't forget the last range
    if current_thresholds is not None:
        ranges.append((current_thresholds, range_start_col, range_end_col))
    
    # Add Average column to the last threshold's range
    if ranges and avg_col:
        last_thresholds, last_start, last_end = ranges[-1]
        ranges[-1] = (last_thresholds, last_start, avg_col)
    
    # Only return if there are multiple ranges (otherwise use static formatting)
    return ranges if len(ranges) > 1 else None


# =============================================================================
# FORMATTING RULES
# =============================================================================

def apply_billing_rules(ws, row_range, first_col, row_idx, thresholds, fills):
    """Apply billing formatting: Red < green_min <= Green <= blue_above < Blue"""
    if not thresholds:
        logging.warning(f"No billing thresholds for row {row_idx}")
        return False
    
    green_min = float(thresholds.get('green_min', 0))
    blue_above = float(thresholds.get('blue_above', 100))
    
    col_letter = get_column_letter(first_col)
    
    ws.conditional_formatting.add(row_range,
        FormulaRule(formula=[f'=LEN(TRIM({col_letter}{row_idx}))=0'], fill=fills['white']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThan', formula=[str(blue_above)], fill=fills['blue']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(green_min)], fill=fills['green']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='lessThan', formula=[str(green_min)], fill=fills['red']))
    
    return True


def apply_ceased_rules(ws, row_range, first_col, row_idx, ceased_thresholds, fills):
    """Apply ceased formatting (inverted): Blue < 0.025 <= Green < 0.04 <= Red"""
    thresholds = ceased_thresholds if ceased_thresholds else DEFAULT_CEASED_THRESHOLDS
    
    blue_below = float(thresholds.get('blue_below', 0.025))
    red_above = float(thresholds.get('red_above', 0.04))
    
    col_letter = get_column_letter(first_col)
    
    ws.conditional_formatting.add(row_range,
        FormulaRule(formula=[f'=LEN(TRIM({col_letter}{row_idx}))=0'], fill=fills['white']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='lessThan', formula=[str(blue_below)], fill=fills['blue']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='lessThan', formula=[str(red_above)], fill=fills['green']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(red_above)], fill=fills['red']))
    
    return True


def apply_rating_rules(ws, row_range, first_col, row_idx, rating_thresholds, fills):
    """Apply 1-5 rating scale: Red(1) -> Amber(2) -> Yellow(3) -> Green(4) -> Blue(5)"""
    thresholds = rating_thresholds if rating_thresholds else DEFAULT_RATING_THRESHOLDS
    
    col_letter = get_column_letter(first_col)
    
    ws.conditional_formatting.add(row_range,
        FormulaRule(formula=[f'=LEN(TRIM({col_letter}{row_idx}))=0'], fill=fills['white']))
    
    threshold_map = {t['rating']: t for t in thresholds}
    
    rating_5_min = float(threshold_map.get(5, {}).get('min', 4.5))
    rating_4_min = float(threshold_map.get(4, {}).get('min', 3.5))
    rating_3_min = float(threshold_map.get(3, {}).get('min', 2.5))
    rating_2_min = float(threshold_map.get(2, {}).get('min', 1.5))
    
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(rating_5_min)], fill=fills['blue']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(rating_4_min)], fill=fills['green']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(rating_3_min)], fill=fills['yellow']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='greaterThanOrEqual', formula=[str(rating_2_min)], fill=fills['amber']))
    ws.conditional_formatting.add(row_range,
        CellIsRule(operator='lessThan', formula=[str(rating_2_min)], fill=fills['red']))
    
    return True


# =============================================================================
# FORMAT AVERAGE TABLES
# =============================================================================

def format_average_table(ws, table_name, table_config, sheet_config, config, fills, year=None):
    """Format average tables - each row may have different KPI type.
    
    For billing rows, supports historical threshold changes via Config_TeamAve_Thresholds.
    """
    table_info = get_table_info(ws, table_name)
    if not table_info or not table_info['data_cols']:
        logging.warning(f"Average table '{table_name}' not found or empty")
        return 0
    
    first_col = min(table_info['data_cols'].values())
    last_col = max(table_info['data_cols'].values())
    team_type = table_config.get('team_type') or sheet_config.get('team_type')
    team_name = sheet_config.get('team_name')  # e.g., 'Physio_North', 'OT'
    
    # For MMP tables, derive team_name from table name
    if sheet_config.get('is_mmp'):
        if 'North' in table_name:
            team_name = 'Physio_North'
        elif 'South' in table_name:
            team_name = 'Physio_South'
        elif 'OT' in table_name:
            team_name = 'OT'
    
    rows_formatted = 0
    
    for row_idx in range(table_info['min_row'] + 1, table_info['max_row'] + 1):
        row_label = ws.cell(row=row_idx, column=table_info['min_col']).value
        if not row_label:
            continue
        row_label = str(row_label).strip()
        row_range = f"{get_column_letter(first_col)}{row_idx}:{get_column_letter(last_col)}{row_idx}"
        
        row_kpi_type = detect_row_kpi_type(row_label)
        
        if row_kpi_type == 'billing':
            # Check for team average threshold history - use column ranges if available
            threshold_ranges = get_team_ave_threshold_ranges(
                team_name, year, table_info['data_cols'], config
            )
            
            if threshold_ranges:
                # Apply different thresholds to different column ranges
                for thresholds, start_col, end_col in threshold_ranges:
                    if thresholds:
                        range_str = f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}"
                        apply_billing_rules(ws, range_str, start_col, row_idx, thresholds, fills)
                rows_formatted += 1
                logging.info(f"  {table_name} billing row: {len(threshold_ranges)} threshold ranges applied")
            else:
                # Fall back to static thresholds
                thresholds = config.get('thresholds', {}).get(team_type, {}).get('Team Average')
                if not thresholds:
                    thresholds = config.get('thresholds', {}).get(team_type, {}).get('CA')
                if thresholds:
                    apply_billing_rules(ws, row_range, first_col, row_idx, thresholds, fills)
                    rows_formatted += 1
                else:
                    logging.warning(f"No Team Average or CA thresholds for {team_type}")
                
        elif row_kpi_type == 'ceased':
            apply_ceased_rules(ws, row_range, first_col, row_idx,
                              config.get('ceased_thresholds', {}), fills)
            rows_formatted += 1
            
        else:
            apply_rating_rules(ws, row_range, first_col, row_idx,
                              config.get('rating_thresholds', []), fills)
            rows_formatted += 1
    
    logging.info(f"Formatted {rows_formatted} rows in average table '{table_name}'")
    return rows_formatted


# =============================================================================
# FORMAT REGULAR TABLES
# =============================================================================

def format_regular_table(ws, table_name, table_config, sheet_config, competency_map, config, fills, year=None):
    """Format regular tables - all rows have same KPI type.
    
    For billing tables with competency history, applies range-based formatting
    with different thresholds for different month ranges.
    """
    table_info = get_table_info(ws, table_name)
    if not table_info or not table_info['data_cols']:
        logging.warning(f"Table '{table_name}' not found or empty")
        return 0
    
    first_col = min(table_info['data_cols'].values())
    last_col = max(table_info['data_cols'].values())
    kpi_type = table_config.get('kpi_type', 'rating')
    team_type = table_config.get('team_type') or sheet_config.get('team_type')
    
    rows_formatted = 0
    
    for row_idx in range(table_info['min_row'] + 1, table_info['max_row'] + 1):
        row_label = ws.cell(row=row_idx, column=table_info['min_col']).value
        if not row_label:
            continue
        row_label = str(row_label).strip()
        
        if row_label not in competency_map:
            logging.debug(f"Skipping '{row_label}' - not in Config_Therapists")
            continue
        
        row_range = f"{get_column_letter(first_col)}{row_idx}:{get_column_letter(last_col)}{row_idx}"
        
        if kpi_type == 'billing':
            # Check for competency history - use column ranges if available
            comp_ranges = get_competency_ranges_for_therapist(
                row_label, year, table_info['data_cols'], config
            )
            
            if comp_ranges:
                # Apply different thresholds to different column ranges
                for competency, start_col, end_col in comp_ranges:
                    thresholds = config.get('thresholds', {}).get(team_type, {}).get(competency)
                    if thresholds:
                        range_str = f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}"
                        apply_billing_rules(ws, range_str, start_col, row_idx, thresholds, fills)
                rows_formatted += 1
                logging.debug(f"  {row_label}: {len(comp_ranges)} competency ranges applied")
            else:
                # Standard single competency for whole row
                competency = competency_map.get(row_label)
                thresholds = config.get('thresholds', {}).get(team_type, {}).get(competency)
                if thresholds:
                    apply_billing_rules(ws, row_range, first_col, row_idx, thresholds, fills)
                    rows_formatted += 1
                else:
                    logging.warning(f"No {team_type}/{competency} thresholds for '{row_label}'")
                
        elif kpi_type == 'ceased':
            apply_ceased_rules(ws, row_range, first_col, row_idx,
                              config.get('ceased_thresholds', {}), fills)
            rows_formatted += 1
            
        elif kpi_type == 'rating':
            apply_rating_rules(ws, row_range, first_col, row_idx,
                              config.get('rating_thresholds', []), fills)
            rows_formatted += 1
    
    logging.info(f"Formatted {rows_formatted} rows in table '{table_name}' ({kpi_type})")
    return rows_formatted


# =============================================================================
# MAIN FUNCTIONS
# =============================================================================

def format_kpi_table(ws, table_name, table_config, sheet_config, competency_map, config, year=None):
    """Dispatch to appropriate formatter based on table type."""
    fills = create_fills(config.get('colours', {}))
    kpi_type = table_config.get('kpi_type', 'rating')
    
    if kpi_type in ('average', 'mmp_average') or table_config.get('is_average'):
        return format_average_table(ws, table_name, table_config, sheet_config, config, fills, year)
    else:
        return format_regular_table(ws, table_name, table_config, sheet_config, competency_map, config, fills, year)


def format_team_leader_sheet(ws, sheet_name, config, year=None):
    """Format all tables in a Team Leader sheet."""
    logging.info(f"Formatting sheet: {sheet_name}")
    
    if sheet_name not in SHEET_CONFIG:
        logging.warning(f"Unknown sheet: {sheet_name}")
        return {'tables_formatted': 0, 'rows_formatted': 0}
    
    sheet_config = SHEET_CONFIG[sheet_name]
    
    clear_all_conditional_formatting(ws)
    
    team_name = sheet_config.get('team_name')
    if sheet_config.get('is_mmp'):
        competency_map = {}
    else:
        competency_map = build_competency_map(config, team_name)
    
    stats = {'tables_formatted': 0, 'rows_formatted': 0}
    
    for table_name, table_config in sheet_config['tables'].items():
        rows = format_kpi_table(ws, table_name, table_config, sheet_config, competency_map, config, year)
        if rows > 0:
            stats['tables_formatted'] += 1
            stats['rows_formatted'] += rows
    
    logging.info(f"Completed {sheet_name}: {stats['tables_formatted']} tables, {stats['rows_formatted']} rows")
    return stats


def format_all_team_leader_sheets(wb, config, file_path=None):
    """Format all Team Leader Dashboard sheets."""
    logging.info("="*60)
    logging.info("Starting Team Leader Dashboard formatting")
    logging.info("="*60)
    
    # Extract year from filename for competency history
    year = extract_year_from_filename(file_path)
    if year:
        logging.info(f"Year extracted from filename: {year}")
    
    # Log competency history status
    history_count = len(config.get('competency_history', []))
    if history_count > 0:
        logging.info(f"Competency history: {history_count} records loaded")
    
    logging.info(f"Ceased thresholds: {config.get('ceased_thresholds', 'USING DEFAULTS')}")
    logging.info(f"Rating thresholds: {len(config.get('rating_thresholds', []))} entries")
    
    total_stats = {'sheets_formatted': 0, 'tables_formatted': 0, 'rows_formatted': 0}
    
    for sheet_name in SHEET_CONFIG.keys():
        if sheet_name not in wb.sheetnames:
            logging.warning(f"Sheet '{sheet_name}' not found - skipping")
            continue
        
        ws = wb[sheet_name]
        stats = format_team_leader_sheet(ws, sheet_name, config, year)
        
        if stats['tables_formatted'] > 0:
            total_stats['sheets_formatted'] += 1
            total_stats['tables_formatted'] += stats['tables_formatted']
            total_stats['rows_formatted'] += stats['rows_formatted']
    
    logging.info("="*60)
    logging.info(f"COMPLETE: {total_stats['sheets_formatted']} sheets, "
                 f"{total_stats['tables_formatted']} tables, "
                 f"{total_stats['rows_formatted']} rows")
    logging.info("="*60)
    
    return total_stats
